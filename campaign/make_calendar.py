import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "30-Day Calendar"

# ── Colors ──────────────────────────────────────────────────────────────────
C_PRIMARY  = "FF009F"
C_SEC      = "FF2D55"
C_ORANGE   = "FF6B4A"
C_YELLOW   = "FFBE3D"
C_GREEN    = "3DB5AA"
C_LIGHT    = "FFF5FB"
C_WHITE    = "FFFFFF"
C_ALT      = "FFE8F7"
C_BLOG     = "FF6B4A"

# ── Header ───────────────────────────────────────────────────────────────────
headers = [
    "Day","Date","Week","Theme","Platform","Content Type",
    "Topic (Arabic)","Hook Line","Design Brief Summary",
    "Image Prompt Summary","File Path","Status"
]
hdr_fill = PatternFill("solid", fgColor=C_PRIMARY)
hdr_font = Font(bold=True, color=C_WHITE, name="Calibri", size=11)
ws.append(headers)
for col, cell in enumerate(ws[1], 1):
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 36

# ── Row Data ──────────────────────────────────────────────────────────────────
start = date(2026, 5, 1)

rows = [
# Day,Week,Theme,Platform,ContentType,TopicAR,HookLine,DesignBrief,ImagePrompt,FilePath
# WEEK 1 — BOLD & PROVOCATIVE
(1,"Week 1","Bold & Provocative","Instagram","Carousel",
 "الماكينة الاقتصادية اللي بتكسب لما تحسي بالنقص",
 "فيه نظام كامل بيكسب لما تحسي إنك ناقصة ☎️",
 "Bold pink gradient; 5 slides exposing the machine; Rubik headlines",
 "Abstract economic gears made of beauty products, pink tones, bold graphic --style bold graphic",
 "campaign/week-01-bold/posts/day-01-carousel.md","Ready to Publish"),

(1,"Week 1","Bold & Provocative","Facebook","Single Post",
 "المرأة العربية تتحكم في 85٪ من قرارات الشراء",
 "85٪ من الفلوس في إيدك — وبرضه بيحسّسوكي إنك ناقصة؟",
 "Stat highlight card; --color-secondary dominant; Rubik bold stat",
 "Large bold 85% numeral on dark pink background, minimalist --style flat illustration",
 "campaign/week-01-bold/posts/day-01-single.md","Ready to Publish"),

(2,"Week 1","Bold & Provocative","Instagram","Story Sequence",
 "إنستجرام وعدم الرضا عن الجسد — الحقيقة اللي بيخبّوها",
 "30 دقيقة على انستجرام كفيلة تخليكي تكرهي جسمك",
 "4-frame story; phone mockup; dark overlay text; --color-primary accents",
 "Phone screen showing filtered vs reality split, soft pink background --style flat illustration",
 "campaign/week-01-bold/stories/day-02-story-sequence.md","Ready to Publish"),

(2,"Week 1","Bold & Provocative","Facebook","Blog Article",
 "الاقتصاد بيكسب لما تحسي بالنقص",
 "مقال: كيف الماكينة الاقتصادية صُمِّمت عشان تحسي إنك ناقصة",
 "Blog cover: bold headline on pink; Rubik font dominant",
 "Woman looking at phone with visible economic chart behind, pink tones --style bold graphic",
 "campaign/blog/article-01-al-iqtisad-wel-nuqsan.md","Ready to Publish"),

(3,"Week 1","Bold & Provocative","Instagram","Infographic",
 "677 مليار دولار — حجم صناعة الجمال العالمية",
 "٦٧٧ مليار دولار — بيكسبوها من إحساسك بالنقص",
 "Infographic; circular stat design; --color-yellow and --color-primary",
 "Infographic style: large dollar sign surrounded by beauty icons, bold yellow and pink palette --style flat illustration",
 "campaign/week-01-bold/infographics/day-03-infographic.md","Ready to Publish"),

(3,"Week 1","Bold & Provocative","Facebook","Story Sequence",
 "كيف الإعلانات بتخلق إحساس بالنقص — تاريخ مخيف",
 "من الخمسينات لحد دلوقتي — نفس الخدعة بوجه مختلف",
 "Timeline story 5 frames; vintage-to-modern aesthetic; --color-orange",
 "Vintage 1950s ad side by side with modern Instagram post, warm tones --style flat illustration",
 "campaign/week-01-bold/stories/day-03-story-sequence.md","Ready to Publish"),

(4,"Week 1","Bold & Provocative","Instagram","Carousel",
 "من الهاوسوايف للسوبروومان — نفس الرسالة القديمة",
 "الإعلانات اتغيّرت.. لكن الفكرة فضلت واحدة",
 "5-slide comparison; split screen layout; --color-secondary and white",
 "Split visual: 1950s housewife vs modern superwoman, bold graphic comparison --style bold graphic",
 "campaign/week-01-bold/posts/day-04-carousel.md","Ready to Publish"),

(4,"Week 1","Bold & Provocative","Facebook","Blog Article",
 "من الهاوسوايف للسوبرومان — نفس الخدعة بشكل مختلف",
 "مقال: التاريخ الإعلاني الكامل وكيف استهدفوا المرأة لـ ٧٠ سنة",
 "Blog: historical timeline cover; mixed vintage-modern design",
 "Collage of advertising eras, women figures through history, editorial style --style flat illustration",
 "campaign/blog/article-04-housewife-to-superwoman.md","Ready to Publish"),

(5,"Week 1","Bold & Provocative","Instagram","Thread",
 "نظرية المقارنة الاجتماعية وتأثيرها على المرأة",
 "ليه لما تدخلي انستجرام تحسي إن حياتك أقل؟ — العلم بيشرح",
 "Thread-style carousel; bullet facts; --color-green contrast",
 "Abstract social comparison visual: two figures, one highlighted, one faded --style flat illustration",
 "campaign/week-01-bold/posts/day-05-thread.md","Ready to Publish"),

(5,"Week 1","Bold & Provocative","Facebook","Single Post",
 "قوة الإنفلونسر في دفع الشراء الاندفاعي",
 "58٪ من قرارات الشراء سببها إنفلونسر — مش صدفة",
 "Stat card; bold quote overlay; --color-orange dominant",
 "Shopping bags emerging from a phone screen, warm orange lighting --style bold graphic",
 "campaign/week-01-bold/posts/day-05-single.md","Ready to Publish"),

(6,"Week 1","Bold & Provocative","Instagram","Story Sequence",
 "صورة الجسد على السوشيال ميديا — تأثير حقيقي",
 "٣٠ دقيقة كفاية تغيّر نظرتك لجسمك — الأبحاث بتثبت",
 "5-frame story: research reveal format; dark overlay; --color-primary",
 "Phone screen with body image content, subtle distortion effect, pink overlay --style bold graphic",
 "campaign/week-01-bold/stories/day-06-story-sequence.md","Ready to Publish"),

(6,"Week 1","Bold & Provocative","Facebook","Infographic",
 "إنفاق الجمال في الشرق الأوسط — 46 مليار دولار",
 "الرقم اللي مش هتصدقيه عن سوق الجمال في منطقتنا",
 "Map infographic; MENA region highlighted; --color-yellow",
 "Middle East map with beauty market statistics overlay, bold yellow tones --style flat illustration",
 "campaign/week-01-bold/infographics/day-06-infographic.md","Ready to Publish"),

(7,"Week 1","Bold & Provocative","Instagram","Carousel",
 "الخوارزمية بتكسب لما تقارني نفسك",
 "البنت المثالية على انستجرام مش حقيقية — دي بيزنس",
 "5-slide expose; algorithm diagram; --color-secondary and --color-light",
 "Algorithm flow chart with beauty content icons, pink and light tones --style bold graphic",
 "campaign/week-01-bold/posts/day-07-carousel.md","Ready to Publish"),

(7,"Week 1","Bold & Provocative","Facebook","Thread",
 "كيف الإعلانات الرقمية تستهدف قلقك",
 "بيعرفوا إنك خايفة من كذا — وبيبيعوا الخوف",
 "Thread text overlay cards; dark pink; Rubik bold",
 "Abstract data points targeting a silhouette, digital advertising visual --style bold graphic",
 "campaign/week-01-bold/posts/day-07-thread.md","Ready to Publish"),

# WEEK 2 — EDUCATE & DATA
(8,"Week 2","Educate & Data","Instagram","Carousel",
 "أرقام لا تصدق: إنفاق الجمال والإعلانات",
 "٧.٧ مليار دولار على إعلانات الجمال في سنة واحدة — حسابي",
 "Data-driven carousel; bar charts; --color-green on dark",
 "Clean data visualization with beauty spend statistics, teal and dark background --style flat illustration",
 "campaign/week-02-educate/posts/day-08-carousel.md","Ready to Publish"),

(8,"Week 2","Educate & Data","Facebook","Blog Article",
 "أرقام لا تصدق: حجم صناعة الجمال والإعلانات",
 "مقال: الأرقام الكاملة عن إنفاق الجمال العالمي والإقليمي",
 "Blog: data-rich infographic cover; teal and pink",
 "Charts and graphs with beauty industry data, clean editorial design --style flat illustration",
 "campaign/blog/article-02-arqam-sinaat-el-gamal.md","Ready to Publish"),

(9,"Week 2","Educate & Data","Instagram","Thread",
 "البحث العلمي عن انستجرام وصورة الجسد",
 "دراسة على ٢٩١ بنت: انستجرام بيخلق مقارنة تصاعدية",
 "Research thread; academic citation style; --color-green",
 "Open research paper with highlighted statistics, clean academic aesthetic --style flat illustration",
 "campaign/week-02-educate/posts/day-09-thread.md","Ready to Publish"),

(9,"Week 2","Educate & Data","Facebook","Story Sequence",
 "تأثير ريلز على القلق النفسي — دراسة عُمان",
 "٣.٢ ساعة يومياً على انستجرام — و٨٠٪ حاسين إنهم بيستخدموه زيادة",
 "4-frame research reveal story; --color-secondary",
 "Phone with reels content, anxiety visualization, minimal pink design --style flat illustration",
 "campaign/week-02-educate/stories/day-09-story-sequence.md","Ready to Publish"),

(10,"Week 2","Educate & Data","Instagram","Infographic",
 "آلية الإنفلونسر ماركتينج في دفع الشراء",
 "من المصداقية لنية الشراء — الخطوات اللي بتمر بيها مش هتعرفيها",
 "Funnel infographic; influencer → trust → purchase; --color-yellow",
 "Purchase funnel diagram with influencer at top, yellow tones --style flat illustration",
 "campaign/week-02-educate/infographics/day-10-infographic.md","Ready to Publish"),

(10,"Week 2","Educate & Data","Facebook","Single Post",
 "نسبة البنات اللي عندهن صورة جسد سلبية في السعودية",
 "٧١٪ من النساء البالغات في السعودية عندهن صورة جسد سلبية",
 "Stat highlight; --color-secondary; bold Rubik; compassionate tone",
 "Soft pink background with bold 71% statistic, gentle lighting --style flat illustration",
 "campaign/week-02-educate/posts/day-10-single.md","Ready to Publish"),

(11,"Week 2","Educate & Data","Instagram","Carousel",
 "اضطراب تشوّه صورة الجسد في المجتمعات العربية",
 "١٣.٥٪ من البنات اللبنانيات عندهن BDD مرتبط بالسوشيال ميديا",
 "Research data carousel; clinical feel softened with pink; Rubik",
 "Abstract mirror reflection showing distorted vs real perception --style flat illustration",
 "campaign/week-02-educate/posts/day-11-carousel.md","Ready to Publish"),

(11,"Week 2","Educate & Data","Facebook","Blog Article",
 "انستجرام وصورة الجسد — الأبحاث بتقول إيه؟",
 "مقال: كل الدراسات العلمية عن انستجرام وصورة الجسد باللغة العربية",
 "Blog: academic meets human; pink and white; diagram cover",
 "Split image: happy woman vs phone screen with filtered content --style flat illustration",
 "campaign/blog/article-03-instagram-wel-gosm.md","Ready to Publish"),

(12,"Week 2","Educate & Data","Instagram","Story Sequence",
 "الضغط الثقافي المضاعف على المرأة العربية",
 "معايير محلية + معايير غربية + معايير السوشيال = ضغط مش طبيعي",
 "5-frame cultural analysis story; --color-orange",
 "Three overlapping cultural layers visual, orange tones --style bold graphic",
 "campaign/week-02-educate/stories/day-12-story-sequence.md","Ready to Publish"),

(12,"Week 2","Educate & Data","Facebook","Single Post",
 "سوق العناية الشخصية في MENA - 95 مليار بحلول 2030",
 "٩٥ مليار دولار بحلول ٢٠٣٠ — ومعظم المنفقين نساء شابات",
 "Market projection card; --color-yellow; upward arrow graphic",
 "Rising graph with MENA map, dollar signs, yellow warm tones --style flat illustration",
 "campaign/week-02-educate/posts/day-12-single.md","Ready to Publish"),

(13,"Week 2","Educate & Data","Instagram","Infographic",
 "كيف الخوارزمية تكثّف المقارنة الاجتماعية",
 "الخوارزمية مش بتعرضلك اللي تحبيه — بتعرضلك اللي يخليكي تفضلي",
 "Algorithm diagram infographic; loop arrows; --color-primary",
 "Circular algorithm diagram with content feed icons, pink and dark --style bold graphic",
 "campaign/week-02-educate/infographics/day-13-infographic.md","Ready to Publish"),

(13,"Week 2","Educate & Data","Facebook","Thread",
 "تاريخ الإعلانات التي استهدفت خوف المرأة",
 "من صابونة الغسيل لبودرة التنحيف — خوفك هو المنتج",
 "Historical ad thread; vintage aesthetic meets modern facts; --color-orange",
 "Vintage advertisement poster with modern data overlay --style flat illustration",
 "campaign/week-02-educate/posts/day-13-thread.md","Ready to Publish"),

(14,"Week 2","Educate & Data","Instagram","Carousel",
 "نظام المقارنة على منصات التواصل — آلية الاستهلاك",
 "كل 'روتين' جديد عند الإنفلونسر معناه منتج ناقص عندك",
 "5-slide data carousel; product → desire loop; --color-green",
 "Product desire loop diagram with influencer content --style flat illustration",
 "campaign/week-02-educate/posts/day-14-carousel.md","Ready to Publish"),

(14,"Week 2","Educate & Data","Facebook","Blog Article",
 "الإنفلونسر ماركتينج وآلية دفعك للشراء",
 "مقال: كيف بالضبط الإنفلونسر بيدفعك تشتري وانت مش واعية",
 "Blog: behind-the-scenes influencer mechanics cover",
 "Influencer filming with purchase icons flowing from phone, editorial style --style bold graphic",
 "campaign/blog/article-05-influencer-marketing.md","Ready to Publish"),

# WEEK 3 — SHIFT
(15,"Week 3","Shift","Instagram","Carousel",
 "الوعي كأداة تحرر — الخطوة الأولى",
 "أول خطوة للتحرر من الماكينة: إنك تشوفيها",
 "Warm shift in tone; sunrise palette; --color-yellow and --color-green",
 "Woman looking at sunrise, warm yellow light, awakening visual --style flat illustration",
 "campaign/week-03-shift/posts/day-15-carousel.md","Ready to Publish"),

(15,"Week 3","Shift","Facebook","Single Post",
 "الفرق بين الاختيار الواعي والاستهلاك الاندفاعي",
 "مش لازم تبطلي تشتري — بس لازم تعرفي ليه بتشتري",
 "Clean comparison card; --color-green; Rubik gentle tone",
 "Two shopping bags: one labelled conscious, one impulsive, gentle green tones --style flat illustration",
 "campaign/week-03-shift/posts/day-15-single.md","Ready to Publish"),

(16,"Week 3","Shift","Instagram","Story Sequence",
 "4 ركائز دكان البنات — طريق الاكتفاء",
 "الاكتفاء الذاتي | الوعي الواعي | الروتين | مجتمع الدعم",
 "4-pillar story sequence; each frame = one pillar; warm tones",
 "Four gentle illustrations for each pillar, warm light palette --style flat illustration",
 "campaign/week-03-shift/stories/day-16-story-sequence.md","Ready to Publish"),

(16,"Week 3","Shift","Facebook","Blog Article",
 "الوعي كأداة تحرر — ليه المعرفة هي القوة الحقيقية",
 "مقال: كيف الوعي وحده يغيّر علاقتك بالاستهلاك",
 "Blog: enlightenment concept cover; warm light; peaceful",
 "Woman reading with warm light, books and nature elements --style flat illustration",
 "campaign/blog/article-06-el-waay-kaodat-tahror.md","Ready to Publish"),

(17,"Week 3","Shift","Instagram","Carousel",
 "Minimalism والجمال الحقيقي — أقل هو أكثر",
 "الجمال الحقيقي مش في كتر المنتجات — في الاختيار الواعي",
 "Minimalist aesthetic; white space; --color-light dominant",
 "Clean minimalist beauty setup, single product, soft natural light --style photorealistic",
 "campaign/week-03-shift/posts/day-17-carousel.md","Ready to Publish"),

(17,"Week 3","Shift","Facebook","Story Sequence",
 "روتين الصباح الواعي — ابدئي يومك بوعي",
 "روتينك الصباحي ممكن يكون مساحة هدوء مش ضغط",
 "Morning routine story; gentle pace; --color-yellow sunrise tone",
 "Peaceful morning routine items, soft yellow sunrise light --style photorealistic",
 "campaign/week-03-shift/stories/day-17-story-sequence.md","Ready to Publish"),

(18,"Week 3","Shift","Instagram","Thread",
 "كيف تبني روتين يومي واعٍ لا استهلاكي",
 "روتين واعي مش معناه تزهدي في نفسك — معناه تعرفي إيه اللي فعلاً محتاجاه",
 "Practical thread; checklist style; --color-green",
 "Clean lifestyle journal with natural elements, teal tones --style photorealistic",
 "campaign/week-03-shift/posts/day-18-thread.md","Ready to Publish"),

(18,"Week 3","Shift","Facebook","Infographic",
 "مقارنة: حياة الاستهلاك vs حياة الوعي",
 "شيفتي من حياة الاستهلاك لحياة الوعي — الفرق بالأرقام",
 "Split infographic; consumption vs awareness; --color-secondary vs --color-green",
 "Before/after lifestyle split infographic, pink vs green palette --style flat illustration",
 "campaign/week-03-shift/infographics/day-18-infographic.md","Ready to Publish"),

(19,"Week 3","Shift","Instagram","Carousel",
 "الاكتفاء الذاتي — الركيزة الأولى",
 "الاكتفاء مش معناه إنك مش محتاجة — معناه إنك كاملة",
 "Empowering warm carousel; --color-yellow and white",
 "Woman standing confidently, warm golden light, natural setting --style photorealistic",
 "campaign/week-03-shift/posts/day-19-carousel.md","Ready to Publish"),

(19,"Week 3","Shift","Facebook","Single Post",
 "Clean Beauty — الجمال النظيف فلسفة حياة",
 "Clean Beauty مش مجرد منتجات خالية من الكيماويات — هي طريقة تفكير",
 "Philosophy card; clean aesthetic; --color-green dominant",
 "Natural beauty ingredients arranged artfully, green and white tones --style photorealistic",
 "campaign/week-03-shift/posts/day-19-single.md","Ready to Publish"),

(20,"Week 3","Shift","Instagram","Story Sequence",
 "مجتمع النساء كقوة — مش أحد ينجح وحده",
 "أقوى سلاح عندك: مجتمع نساء بيدعمك",
 "Community story; connected women silhouettes; --color-primary warm",
 "Interconnected women silhouettes, warm pink community visual --style flat illustration",
 "campaign/week-03-shift/stories/day-20-story-sequence.md","Ready to Publish"),

(20,"Week 3","Shift","Facebook","Blog Article",
 "Minimalism والجمال الحقيقي — فلسفة العيش الواعي",
 "مقال: الـ Minimalism مش تقتير — هي حرية",
 "Blog: minimalist cover; white space; single bold element",
 "Single flower in clean white vase, natural light, minimal composition --style photorealistic",
 "campaign/blog/article-07-minimalism-wel-gamal.md","Ready to Publish"),

(21,"Week 3","Shift","Instagram","Infographic",
 "الأعمدة الأربعة لحياة واعية — خريطة كاملة",
 "٤ أعمدة تبنيكي من الداخل — مش من الخارج",
 "4-pillar diagram; clean; --color-green and --color-yellow",
 "Clean four-pillar infographic diagram, teal and yellow, minimal icons --style flat illustration",
 "campaign/week-03-shift/infographics/day-21-infographic.md","Ready to Publish"),

(21,"Week 3","Shift","Facebook","Thread",
 "لماذا الانفصال عن معايير الجمال المعولمة ضروري",
 "معايير جمال غربية + شرق آسيوية + محلية = توقعات مستحيلة",
 "Thread: cultural deconstruction; --color-orange warm",
 "Multiple cultural beauty standards as overlapping circles, orange warm tones --style bold graphic",
 "campaign/week-03-shift/posts/day-21-thread.md","Ready to Publish"),

# WEEK 4 — EMPOWER
(22,"Week 4","Empower","Instagram","Carousel",
 "أنتِ كاملة — الوعي هو الفرق",
 "مش ناقصة منتج — ناقصة وعي بقيمتك الحقيقية",
 "Celebratory carousel; gold and pink; Rubik bold empowerment",
 "Woman radiant and confident, golden hour light, empowerment visual --style photorealistic",
 "campaign/week-04-empower/posts/day-22-carousel.md","Ready to Publish"),

(22,"Week 4","Empower","Facebook","Story Sequence",
 "رحلة إعادة ضبط العقلية — من الإحساس بالنقص للاكتمال",
 "من يوم ١ للنهاية — شوفي كام مشيتي",
 "Journey recap story; 5 frames; all brand colors",
 "Journey path visualization, warm glow at destination, pink and gold --style flat illustration",
 "campaign/week-04-empower/stories/day-22-story-sequence.md","Ready to Publish"),

(23,"Week 4","Empower","Instagram","Thread",
 "مجتمع دكان البنات — مين هي البنت اللي بنبنيها معاها",
 "بنت الـ Ecosystem: واعية، طموحة، حقيقية",
 "Community portrait thread; warm; --color-primary",
 "Diverse Arab women, warm community setting, authentic expressions --style photorealistic",
 "campaign/week-04-empower/posts/day-23-thread.md","Ready to Publish"),

(23,"Week 4","Empower","Facebook","Blog Article",
 "مجتمع النساء كقوة — لماذا لا أحد ينجح وحده",
 "مقال: قوة مجتمع النساء وكيف دكان البنات بيبنيه",
 "Blog: warm community cover; women together; pink",
 "Women in supportive circle, warm golden light, community feeling --style photorealistic",
 "campaign/blog/article-09-mogtamaa-al-nissa.md","Ready to Publish"),

(24,"Week 4","Empower","Instagram","Carousel",
 "صاحبة المشروع الواعية — من الفكرة للبراند",
 "فكرتك + وعيك = براند حقيقي — إزاي؟",
 "Entrepreneurship carousel; step-by-step; --color-orange",
 "Woman building brand identity, creative workspace, warm orange tones --style photorealistic",
 "campaign/week-04-empower/posts/day-24-carousel.md","Ready to Publish"),

(24,"Week 4","Empower","Facebook","Infographic",
 "نظام دكان البنات — Ecosystem متكامل",
 "من الجمال الواعي للبيزنس الحقيقي — كل حاجة في مكان واحد",
 "Ecosystem diagram infographic; all 6 pillars; --color-primary",
 "Six-pillar ecosystem diagram, pink and teal, icons for each service --style flat illustration",
 "campaign/week-04-empower/infographics/day-24-infographic.md","Ready to Publish"),

(25,"Week 4","Empower","Instagram","Thread",
 "Clean Marketing — التسويق النظيف والثقة الحقيقية",
 "التسويق النظيف: تبيعي من غير ما تستغلي خوف حد",
 "Marketing ethics thread; clean aesthetic; --color-green",
 "Clean marketing concept: genuine handshake between brand and customer --style flat illustration",
 "campaign/week-04-empower/posts/day-25-thread.md","Ready to Publish"),

(25,"Week 4","Empower","Facebook","Blog Article",
 "Clean Marketing — التسويق النظيف والثقة الحقيقية",
 "مقال: إزاي تعملي بيزنس ناجح من غير ما تستغلي خوف العميلة",
 "Blog: clean professional cover; trust symbols; teal",
 "Clean workspace with brand trust elements, teal and white --style photorealistic",
 "campaign/blog/article-11-clean-marketing.md","Ready to Publish"),

(26,"Week 4","Empower","Instagram","Carousel",
 "قصص نجاح بنات الـ Ecosystem",
 "هي بدأت زيك — وعيها هو اللي خلّى مشروعها يكبر",
 "Success story carousel; real community feel; --color-yellow",
 "Woman entrepreneur celebrating success, warm yellow light, authentic feel --style photorealistic",
 "campaign/week-04-empower/posts/day-26-carousel.md","Ready to Publish"),

(26,"Week 4","Empower","Facebook","Story Sequence",
 "كيف تبني براندك بوعي على دكان البنات",
 "خطوة خطوة: من الفكرة لمتجر حقيقي على دكان البنات",
 "Step-by-step story; 5 frames; --color-primary",
 "Brand building steps visual, pink gradient, icons for each step --style flat illustration",
 "campaign/week-04-empower/stories/day-26-story-sequence.md","Ready to Publish"),

(27,"Week 4","Empower","Instagram","Story Sequence",
 "اللحظة اللي تقرري فيها إنك كاملة — رسالة للجمهور",
 "انتِ كاملة — مش بعد المنتج الجديد — دلوقتي",
 "Emotional story; woman journey visual; --color-primary warm",
 "Woman peaceful moment, soft pink light, self-acceptance visual --style photorealistic",
 "campaign/week-04-empower/stories/day-27-story-sequence.md","Ready to Publish"),

(27,"Week 4","Empower","Facebook","Thread",
 "برنامج الـ Loyalty في دكان البنات — كيف تستفيدي",
 "مش بس متابعة — اتعملي عضوة وابني مكانك في المجتمع",
 "Loyalty program thread; membership feel; --color-primary",
 "Loyalty card visual, member benefits icons, pink elegant design --style flat illustration",
 "campaign/week-04-empower/posts/day-27-thread.md","Ready to Publish"),

(28,"Week 4","Empower","Instagram","Carousel",
 "وعي الاستهلاك — الدليل العملي النهائي",
 "١٠ أسئلة تسأليها نفسك قبل أي شراء — دليلك الواعي",
 "Practical guide carousel; checklist style; --color-green",
 "Clean checklist visual, teal tones, practical everyday items --style flat illustration",
 "campaign/week-04-empower/posts/day-28-carousel.md","Ready to Publish"),

(28,"Week 4","Empower","Facebook","Blog Article",
 "كيف تبني روتين يومي قائم على الوعي لا الاستهلاك",
 "مقال: دليل الروتين اليومي الواعي — خطوة خطوة",
 "Blog: morning routine cover; warm light; practical",
 "Morning routine flat lay, natural elements, warm morning light --style photorealistic",
 "campaign/blog/article-08-routine-yawmi-waee.md","Ready to Publish"),

(29,"Week 4","Empower","Instagram","Thread",
 "رسالة لكل بنت عربية طموحة — أنتِ ممكن",
 "مش بس ممكن — أنتِ بالفعل كل اللي تحتاجيه موجود فيكِ",
 "Empowerment manifesto thread; bold; --color-primary",
 "Woman standing strong, golden light behind, empowerment composition --style photorealistic",
 "campaign/week-04-empower/posts/day-29-thread.md","Ready to Publish"),

(29,"Week 4","Empower","Facebook","Infographic",
 "ملخص الـ 30 يوم — رحلة Reset Your Mindset",
 "٣٠ يوم، ٤ مراحل، رسالة واحدة: أنتِ كاملة",
 "Campaign summary infographic; all phases; --color-primary full",
 "Journey summary infographic: four phases with icons, pink and teal --style flat illustration",
 "campaign/week-04-empower/infographics/day-29-infographic.md","Ready to Publish"),

(30,"Week 4","Empower","Instagram","Carousel",
 "Reset Your Mindset — الدليل الكامل",
 "رحلتك مع دكان البنات بدأت — والجزء الأجمل لسه جاي",
 "Grand finale carousel; all brand colors celebration",
 "Celebration visual: woman reaching upward, pink confetti, golden light --style photorealistic",
 "campaign/week-04-empower/posts/day-30-carousel.md","Ready to Publish"),

(30,"Week 4","Empower","Facebook","Story Sequence",
 "ادخلي دكان البنات — مجتمعك ينتظرك",
 "مجتمع حقيقي، محتوى واعٍ، بيزنس حقيقي — كلها في مكانك الجديد",
 "Final CTA story; 4 frames; all colors; dokanelbanat.com",
 "Community gathering visual, warm celebratory feel, pink and gold --style photorealistic",
 "campaign/week-04-empower/stories/day-30-story-sequence.md","Ready to Publish"),

(30,"Week 4","Empower","Instagram+Facebook","Blog Article",
 "Reset Your Mindset — الدليل الكامل لإعادة ضبط العقلية",
 "مقال: كل حاجة تحتاجي تعرفيها عشان تبدئي Reset Your Mindset",
 "Blog: flagship pillar article cover; bold and inspiring",
 "Woman in powerful pose, sunrise background, reset concept visual --style photorealistic",
 "campaign/blog/article-12-reset-your-mindset-guide.md","Ready to Publish"),
]

# ── Write rows ───────────────────────────────────────────────────────────────
alt_fill  = PatternFill("solid", fgColor=C_ALT)
base_fill = PatternFill("solid", fgColor="FFFFFF")

week_fills = {
    "Week 1": PatternFill("solid", fgColor="FFF0FA"),
    "Week 2": PatternFill("solid", fgColor="F0FBF9"),
    "Week 3": PatternFill("solid", fgColor="FFFBF0"),
    "Week 4": PatternFill("solid", fgColor="FFF5F0"),
}

for i, r in enumerate(rows, 2):
    day, wk, theme, platform, ct, topic, hook, design, imgprompt, fp, _status = r
    row_date = start + timedelta(days=day - 1)
    ws.append([day, row_date.strftime("%Y-%m-%d"), wk, theme, platform,
               ct, topic, hook, design, imgprompt, fp, "Ready to Publish"])
    fill = week_fills.get(wk, base_fill)
    for col in range(1, 13):
        cell = ws.cell(row=i, column=col)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 60

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [6, 12, 8, 22, 20, 18, 40, 45, 50, 55, 55, 18]
for col_i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_i)].width = w

ws.freeze_panes = "A2"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.append(["Content Type","Count"])
from collections import Counter
ct_counts = Counter(r[4] for r in rows)
for ct, cnt in sorted(ct_counts.items()):
    ws2.append([ct, cnt])
ws2.append([])
ws2.append(["Platform","Count"])
plat_counts = Counter(r[3] for r in rows)
for p, cnt in sorted(plat_counts.items()):
    ws2.append([p, cnt])

# ── Save ──────────────────────────────────────────────────────────────────────
path = r"D:\claude-Projects\dokanelbanat\campaign\calendar.xlsx"
wb.save(path)
print(f"Saved: {path}")
import os
print(f"Size: {os.path.getsize(path):,} bytes")
print("CALENDAR DONE")

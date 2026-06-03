#!/usr/bin/env python3
"""
Bootstrap + build script for dokanelbanat.com calendar.
Auto-installs openpyxl if missing, then builds calendar.xlsx.
Run with: python bootstrap_and_build.py
"""
import sys, subprocess, os

print(f"Python {sys.version}")

# ── Auto-install openpyxl if needed ──────────────────────────────────────────
try:
    import openpyxl
    print(f"openpyxl {openpyxl.__version__} already available.")
except ImportError:
    print("openpyxl not found — installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    print(f"openpyxl {openpyxl.__version__} installed successfully.")

# ── Now import everything we need ─────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pathlib

OUTPUT_PATH = r"D:\claude-Projects\dokanelbanat\campaign\calendar.xlsx"
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# ── Brand Palette ─────────────────────────────────────────────────────────────
C_PRIMARY   = "FF009F"
C_WHITE     = "FFFFFF"
C_ALT       = "FFF0F8"

# ── Column Headers ────────────────────────────────────────────────────────────
HEADERS = [
    "Day", "Date", "Week", "Theme",
    "Platform", "Content Type", "Topic (Arabic)",
    "Hook Line", "Design Brief Summary",
    "Image Prompt Summary", "File Path", "Status"
]

# ── Full 46-Row Calendar Data ─────────────────────────────────────────────────
# (Day, Date, Week, Theme, Platform, ContentType, TopicArabic,
#  HookLine, DesignBrief, ImagePrompt, FilePath)
ROWS = [

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 1 — BOLD & PROVOCATIVE  |  Days 1-7  |  May 1-7, 2026
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 1 ───────────────────────────────────────────────────────────────────
    (1,"2026-05-01","Week 1 – Bold","Expose the Machine",
     "Instagram","Carousel",
     "ماكينة الإنفاق العالمية على المرأة",
     "كل ما تحسي إنك ناقصة — في حد بيكسب منها.",
     "6-slide carousel. Slide 1: bold #ff009f headline on near-black bg. Slides 2-5: one data stat per slide — large numeral in #ffbe3d, icon, Arabic explanation. Slide 6: dokanelbanat.com URL + follow CTA. Rubik Bold headlines, MarkaziText sub-copy.",
     "Giant chrome industrial machine; conveyor belt feeds shopping bags into a woman silhouette. Neon pink (#ff009f) lighting on deep black. Editorial, high-contrast.",
     "campaign/week-01-bold/posts/day-01-carousel.md"),

    (1,"2026-05-01","Week 1 – Bold","Expose the Machine",
     "Facebook","Single Post",
     "لماذا تشعرين دائماً بالنقص؟",
     "السؤال اللي محدش بيسأله — ليه إحساس النقص ده مش بيروح أبداً؟",
     "Single square image. Left half: chaotic consumerist collage (pink neon). Right half: calm Arab woman looking away. Bold Arabic question spans both halves. #ff009f strip down center. Facebook caption 150 words.",
     "Split-frame: left = chaotic pink-neon shopping mall interior; right = serene Arab woman in natural light looking to horizon. Sharp split line with brand pink as divider. Cinematic ratio.",
     "campaign/week-01-bold/posts/day-01-single.md"),

    # DAY 2 ───────────────────────────────────────────────────────────────────
    (2,"2026-05-02","Week 1 – Bold","Expose the Machine",
     "Instagram","Story Sequence",
     "انستجرام وصورة الجسم عند المرأة العربية",
     "دقيقتين على انستجرام = تغيير في صورتك عن نفسك.",
     "5-story vertical sequence (1080x1920). Story 1: shocking stat on #ff009f full-bleed. Stories 2-3: scrolling stat cards with progress bar. Story 4: poll sticker 'هل حصل معاكي؟'. Story 5: link-sticker to blog article. Gradient transitions.",
     "Phone screen mockup showing Instagram filtered-body feed; crack-glitch effect spreading from screen. Deep pink and charcoal tones. Close-up perspective.",
     "campaign/week-01-bold/stories/day-02-story-sequence.md"),

    (2,"2026-05-02","Week 1 – Bold","Expose the Machine",
     "Blog","Blog Article",
     "انستجرام وأزمة صورة الجسم عند المرأة العربية",
     "كيف تحوّل انستجرام المرأة العربية إلى زبونة دائمة؟",
     "1800-word deep-dive. Sections: research overview, psychological triggers, Arab-specific data, algorithm exploitation of insecurity, dokanelbanat.com alternative. Internal links + CTA.",
     "Hero illustration: Arab woman holding phone; mirror reflection shows idealised filtered body. Painterly style. #ff009f dominant, cream accents. Wide-format blog hero.",
     "campaign/blog/article-01-instagram-body-image.md"),

    # DAY 3 ───────────────────────────────────────────────────────────────────
    (3,"2026-05-03","Week 1 – Bold","Expose the Machine",
     "Instagram","Infographic",
     "صناعة الجمال: 677 مليار دولار — وكيف بُنيت على النقص",
     "$677,000,000,000 — ده حجم الصناعة اللي بتتبنى على إحساسك بالنقص.",
     "Tall portrait infographic (1080x1920). Decade-by-decade timeline 1920s-2026: revenue figure per era + key manipulation tactic. Bold numerals in #ffbe3d on dark charcoal. dokanelbanat.com watermark footer.",
     "Vertical timeline poster: each era has miniature vintage-to-modern ad style and a glowing golden revenue figure. Dark charcoal background, pink timeline spine, decade labels white.",
     "campaign/week-01-bold/infographics/day-03-infographic.md"),

    (3,"2026-05-03","Week 1 – Bold","Expose the Machine",
     "Facebook","Story Sequence",
     "تاريخ إعلانات المرأة ربة المنزل",
     "من 'ابشري يا ست البيت' لـ'بشرتك مش كفاية' — الإعلان دايماً عارف يضغط على الجرح.",
     "4-story Facebook sequence. Story 1: vintage housewife ad aesthetic, sepia overlay. Stories 2-3: decade jump cards (50s to 80s to 2020s) showing evolved manipulation. Story 4: poll 'شايفة نفسك في الإعلانات؟'",
     "Collage of vintage Arabic housewife ads transitioning via film-burn to contemporary social media beauty ads. Sepia to neon pink color progression.",
     "campaign/week-01-bold/stories/day-03-story-sequence.md"),

    # DAY 4 ───────────────────────────────────────────────────────────────────
    (4,"2026-05-04","Week 1 – Bold","Expose the Machine",
     "Instagram","Carousel",
     "سيكولوجية الشراء عند الإنفلوانسر",
     "مش بتشتري المنتج — بتشتري النسخة اللي إنتي عايزة تبقيها.",
     "7-slide carousel. Slide 1: influencer lifestyle still with overlay. Slides 2-6: one psychological trigger per slide (FOMO, social proof, aspirational identity, scarcity, parasocial bond) — trigger name in #ff009f, explanation, real Arab example. Slide 7: awareness CTA.",
     "Stylised illustration: Arab woman gazing at glowing phone; reflection in screen is a glamourised version of herself. Neon pink dominant, dark surround, dream-sequence atmosphere.",
     "campaign/week-01-bold/posts/day-04-carousel.md"),

    (4,"2026-05-04","Week 1 – Bold","Expose the Machine",
     "Blog","Blog Article",
     "سيكولوجية الشراء: لماذا نصدق الإنفلوانسر؟",
     "العلم وراء زر 'اشتري دلوقتي' وليه دماغك مش قادرة تقاومه.",
     "2000-word article: parasocial relationships, aspirational identity consumption, dopamine-reward loops, FOMO mechanics, FTC disclosure failures in Arab markets, Arab influencer case studies. Awareness section at close.",
     "Hero: extreme close-up of Arab woman's eye reflecting glowing 'Buy Now' button. Cinematic lighting, shallow depth of field. Pink and gold tones, psychological tension.",
     "campaign/blog/article-02-influencer-psychology.md"),

    # DAY 5 ───────────────────────────────────────────────────────────────────
    (5,"2026-05-05","Week 1 – Bold","Expose the Machine",
     "Instagram","Thread",
     "نظرية المقارنة الاجتماعية وتأثيرها على المرأة العربية",
     "ليه ما تقدريش تبطلي تقارني نفسك — وده مش عيب فيكي.",
     "8-post thread. Posts: Festinger 1954 theory intro, upward vs downward comparison, social media amplification, body image data, self-esteem impact, Arab cultural layer (family/community), breaking the cycle, dokanelbanat.com as counter-space. Consistent card design.",
     "8 minimal cards: Arabic key phrase centred on clean white/blush. Subtle geometric pattern. Quote or stat in #ff009f beneath each phrase. Unified thread visual identity.",
     "campaign/week-01-bold/posts/day-05-thread.md"),

    (5,"2026-05-05","Week 1 – Bold","Expose the Machine",
     "Facebook","Single Post",
     "المقارنة الاجتماعية في السياق العربي",
     "المجتمع بيعلمنا نقارن — والسوشيال ميديا بتتاجر فيها.",
     "Single educational Facebook post with stats overlay on warm background. Long-form caption (200 words) on how Arab cultural comparison norms (family gatherings, weddings) are turbo-charged by social platforms.",
     "Two Arab women side-by-side (illustrated): one showing phone to other. Phone screen glows with comparison imagery. Warm orange (#ff6b4a) tones. Editorial flat illustration.",
     "campaign/week-01-bold/posts/day-05-single.md"),

    # DAY 6 ───────────────────────────────────────────────────────────────────
    (6,"2026-05-06","Week 1 – Bold","Expose the Machine",
     "Instagram","Story Sequence",
     "إنفاق المرأة العربية على منتجات التجميل",
     "إيه اللي بتنفقيه فعلاً — وعلى إيه بالظبط؟",
     "5-story sequence. Story 1: shocking monthly spend stat full-bleed. Stories 2-3: category breakdown (skincare/makeup/hair/diet). Story 4: interactive quiz 'كام صرفتي الشهر ده على التجميل؟'. Story 5: awareness reframe + dokanelbanat.com link.",
     "Animated-feel story frames: gold coins and beauty products floating on #ff009f background. Arabic numerals bold white. Calculator emoji integrated. Clean energetic composition.",
     "campaign/week-01-bold/stories/day-06-story-sequence.md"),

    (6,"2026-05-06","Week 1 – Bold","Expose the Machine",
     "Instagram","Infographic",
     "أين تذهب أموال المرأة العربية؟",
     "خريطة الإنفاق — من جيبك لجيب مين؟",
     "Horizontal spending-breakdown infographic. Pie chart (centre) + category legend: skincare 35%, makeup 25%, hair 18%, diet & supplements 12%, fragrance 10%. Each slice in distinct brand palette colour. Arabic labels, data source footnote, dokanelbanat.com footer.",
     "Overhead flat-lay arranged as pie chart: beauty products physically grouped by category, each on a color-coded section. White marble base, top-down shot, natural lighting.",
     "campaign/week-01-bold/infographics/day-06-infographic.md"),

    # DAY 7 ───────────────────────────────────────────────────────────────────
    (7,"2026-05-07","Week 1 – Bold","Expose the Machine",
     "Instagram","Carousel",
     "ملخص الأسبوع الأول: الماكينة انكشفت",
     "أسبوع كامل من الحقايق — إيه اللي صدمك أكتر؟",
     "6-slide week-recap carousel. Slide 1: bold 'Week 1 Recap' title on #ff009f full-bleed. Slides 2-5: one key takeaway per topic day. Slide 6: Week 2 teaser 'الأسبوع الجاي: الأرقام والبيانات' + follow CTA.",
     "Mosaic grid of Week 1 post mini-thumbnails in 2x3 layout. WEEK 1 RECAP headline overlaid. #ff009f dominant. Rubik Bold throughout.",
     "campaign/week-01-bold/posts/day-07-carousel.md"),

    (7,"2026-05-07","Week 1 – Bold","Expose the Machine",
     "Facebook","Thread",
     "أسئلة الأسبوع الأول من المجتمع",
     "إنتوا سألتوا — دي إجاباتنا. الحوار الحقيقي بيبدأ هنا.",
     "Facebook thread: root post asks community for their biggest Week 1 question. Top 5 pre-prepared answer comments pinned in order. Each answer cites specific Week 1 post/stat. Engagement loop design.",
     "Community conversation design: overlapping speech bubbles in brand colours with Arabic text inside. Question marks and lightbulb icons. Clean flat illustration on white.",
     "campaign/week-01-bold/posts/day-07-thread.md"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 2 — EDUCATE & DATA  |  Days 8-14  |  May 8-14, 2026
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 8 ───────────────────────────────────────────────────────────────────
    (8,"2026-05-08","Week 2 – Educate","Research & Data",
     "Instagram","Carousel",
     "أحدث أبحاث تأثير السوشيال ميديا على المرأة",
     "مش رأي — ده بحث علمي. الأرقام بتتكلم.",
     "8-slide data carousel. Each slide: source name top-left in MarkaziText, finding headline Rubik Bold #ff009f, one-line explanation, sample size stat. #3db5aa credibility accent bar on left edge. White card on light-grey background.",
     "Clean research-poster aesthetic: white cards on subtle grid paper. Green accent bars, Arabic source citations, data bar charts in brand palette. Academic meets accessible.",
     "campaign/week-02-educate/posts/day-08-carousel.md"),

    (8,"2026-05-08","Week 2 – Educate","Research & Data",
     "Blog","Blog Article",
     "الأبحاث العلمية وتأثير السوشيال ميديا على صورة المرأة",
     "ما تقوله الأبحاث عن السوشيال ميديا وصورة المرأة — والأرقام مفاجئة.",
     "2200-word research synthesis. 5 peer-reviewed studies: body image deterioration, self-esteem correlation, purchase behaviour, social comparison amplification, addiction-loop design. Full Arabic citations, expert quotes.",
     "Hero illustration: academic research papers with social media icons overlaid as bookmarks. Green and pink palette. Bookish but modern editorial style.",
     "campaign/blog/article-03-research-social-media.md"),

    # DAY 9 ───────────────────────────────────────────────────────────────────
    (9,"2026-05-09","Week 2 – Educate","Research & Data",
     "Instagram","Thread",
     "آليات الإنفلوانسر: كيف يعمل النظام من الداخل؟",
     "من العقد للمحتوى للضغط النفسي — رحلة المنتج من المصنع لعقلك.",
     "7-post thread: brand-influencer contract structure, disclosure laws vs. Arab reality, manufactured authenticity, micro vs. macro ROI, relatability illusion mechanics, audience data sold to brands, awareness tools for consumers.",
     "Flowchart-series thread cards: each has process arrow or system diagram showing one influencer machine step. Money-flow arrows, brand boxes, audience funnel icons. Green + pink accent.",
     "campaign/week-02-educate/posts/day-09-thread.md"),

    (9,"2026-05-09","Week 2 – Educate","Research & Data",
     "Facebook","Story Sequence",
     "خلف الكواليس: كيف تختار الشركات الإنفلوانسر؟",
     "العملية اللي ما بتشوفيهاش — وكيف تأثر عليكي.",
     "4-story behind-the-scenes. Story 1: brand brief document mockup. Story 2: engagement-rate spreadsheet visual. Story 3: content approval process diagram. Story 4: poll 'هل تعرفتي على ده من قبل؟'",
     "Documentary-style mockups: contract papers, spreadsheets, brand briefs layered with Instagram grid previews. Moody blue-pink lighting. Newspaper-print meets social-media aesthetic.",
     "campaign/week-02-educate/stories/day-09-story-sequence.md"),

    # DAY 10 ──────────────────────────────────────────────────────────────────
    (10,"2026-05-10","Week 2 – Educate","Research & Data",
      "Instagram","Infographic",
      "أنماط الإعلان التاريخية التي استهدفت المرأة: 1950-2026",
      "من الخمسينيات للألفية — الإعلان دايماً عارف ضعفك.",
      "Tall timeline infographic. 7 eras: 1950s housewife perfection, 1970s liberation paradox, 1980s supermodel, 1990s heroin chic, 2000s celebrity skin, 2010s natural beauty fakeout, 2020s body positivity weaponised. Each: mini ad mockup + manipulation tactic.",
      "Vertical vintage-to-modern timeline collage: top is sepia newspaper aesthetic, bottom is neon digital. Each era has representative ad mockup. Brand pink spine full height.",
      "campaign/week-02-educate/infographics/day-10-infographic.md"),

    (10,"2026-05-10","Week 2 – Educate","Research & Data",
      "Facebook","Single Post",
      "الإعلان التاريخي وتأثيره على المرأة العربية",
      "إزاي السوق علّم المرأة العربية تكره جسمها — وإمتى بدأ ده؟",
      "Single Facebook post with vintage Arabic magazine ad reimagined in modern graphic design. Long-form caption (200+ words) on advertising's historic role in Arab women's body standards. Source citations included.",
      "Split composition: sepia vintage Arabic beauty ad (left) mirrored by equivalent contemporary social media ad (right). Same manipulation message visible across eras.",
      "campaign/week-02-educate/posts/day-10-single.md"),

    # DAY 11 ──────────────────────────────────────────────────────────────────
    (11,"2026-05-11","Week 2 – Educate","Research & Data",
      "Instagram","Carousel",
      "سوق الجمال في منطقة MENA: 46 مليار دولار",
      "46 مليار دولار — ومصر والسعودية في القلب.",
      "6-slide market-data carousel. Slides: MENA total market hero stat, Egypt country breakdown, Saudi breakdown, fastest-growing segments (skincare +18% YoY), key players + emerging Arab-owned brands, power reframe for Arab women consumers.",
      "MENA map with glowing data-point clusters over Egypt, Saudi, UAE. Bar charts in brand colours. Financial-report aesthetic with Arabic editorial typography. Night-map satellite style.",
      "campaign/week-02-educate/posts/day-11-carousel.md"),

    (11,"2026-05-11","Week 2 – Educate","Research & Data",
      "Blog","Blog Article",
      "سوق الجمال في الشرق الأوسط: 46 مليار دولار وأنتِ في المنتصف",
      "من يستفيد فعلاً من سوق الجمال في الشرق الأوسط؟",
      "2000-word market analysis. MENA beauty market structure, controlling multinationals, Arab-owned emerging brands, Egypt and Saudi as key markets, Arab women as consumers vs. potential entrepreneurs, dokanelbanat.com as community counter-model.",
      "Hero: stylised MENA map constructed from beauty product icons. Glowing green and pink overlay. Data-art editorial quality.",
      "campaign/blog/article-04-mena-beauty-market.md"),

    # DAY 12 ──────────────────────────────────────────────────────────────────
    (12,"2026-05-12","Week 2 – Educate","Research & Data",
      "Instagram","Story Sequence",
      "إحصائيات تشوه صورة الجسم (Body Dysmorphia) بين النساء",
      "الأرقام اللي محتاجة تعرفيها — ومش الصحافة بتتكلم فيها.",
      "5-story educational sequence. Story 1: global BDD prevalence stat full-bleed. Story 2: Arab-region data. Story 3: social media correlation research. Story 4: age-of-onset statistics. Story 5: gentle CTA — dokanelbanat.com awareness resources + 'إنتي مش لوحدك'.",
      "Clinical-meets-human: soft watercolour woman looking into fractured mirror; statistics as bold overlaid text. Green (#3db5aa) and pink (#ff009f). Compassionate, not alarming.",
      "campaign/week-02-educate/stories/day-12-story-sequence.md"),

    (12,"2026-05-12","Week 2 – Educate","Research & Data",
      "Facebook","Carousel",
      "تشوه صورة الجسم والشراء القهري: الصلة الخفية",
      "ما بين الوسواس والتسوق — الصلة اللي العلم أثبتها.",
      "5-slide carousel connecting BDD stats to compulsive buying disorder. Cites peer-reviewed sources. Sensitive tone: factual, non-stigmatising. Each slide: research finding + real-life manifestation + awareness note.",
      "Compassionate data mosaic: gentle illustrated faces with embedded statistic numbers, pastel base with #ff009f accents. Informational and warm simultaneously.",
      "campaign/week-02-educate/posts/day-12-carousel.md"),

    # DAY 13 ──────────────────────────────────────────────────────────────────
    (13,"2026-05-13","Week 2 – Educate","Research & Data",
      "Instagram","Infographic",
      "بحث إدمان السوشيال ميديا: الخوارزمية صُممت ضدك",
      "الخوارزمية مش بتوصيلك محتوى — بتوصيلك إحساس.",
      "Tall portrait infographic. 5 sections: Dopamine Loop diagram, Variable Reward Schedule (slot-machine comparison), Infinite Scroll design intent, Comparison-content amplification data, Attention economy dollar value per user. Each cites research source.",
      "Brain anatomy illustration: social media platform logos as neural-pathway nodes; dopamine molecules as pink bubbles. Deep dark background, neon pink and green accents. Scientific editorial.",
      "campaign/week-02-educate/infographics/day-13-infographic.md"),

    (13,"2026-05-13","Week 2 – Educate","Research & Data",
      "Facebook","Thread",
      "إدمان السوشيال ميديا: هل أنتِ مدمنة؟ 7 أسئلة للتقييم",
      "7 أسئلة تحددي بيهم علاقتك الحقيقية مع السوشيال ميديا.",
      "Facebook thread: root post introduces 7-question self-assessment. Each question posted as separate comment with 2-sentence explanation. Final comment: free PDF from dokanelbanat.com with full assessment + scoring guide.",
      "Assessment worksheet aesthetic: question card design, bold Arabic numbers, #ff009f circle bullets, clean white background. Clinical but warm. Shareable as screenshot.",
      "campaign/week-02-educate/posts/day-13-thread.md"),

    # DAY 14 ──────────────────────────────────────────────────────────────────
    (14,"2026-05-14","Week 2 – Educate","Research & Data",
      "Instagram","Carousel",
      "دراسة السعودية: صورة الجسم والشبكات الاجتماعية عند المرأة العربية",
      "دراسة سعودية تكشف: الفتيات العربيات أكتر تأثراً — وده السبب.",
      "7-slide carousel. Slide 1: study headline and key finding. Slides 2-5: Arab-specific cultural factors — modesty paradox, family comparison dynamics, wedding body pressure, skin-lightening demand. Slide 6: Egypt comparison data. Slide 7: Week 2 recap + Week 3 teaser.",
      "Research-presentation style: Saudi cultural imagery (modest dress, family setting) intersected with data charts. Warm tan and #ff009f palette. Academic but culturally grounded.",
      "campaign/week-02-educate/posts/day-14-carousel.md"),

    (14,"2026-05-14","Week 2 – Educate","Research & Data",
      "Blog","Blog Article",
      "الدراسة السعودية: كيف تؤثر وسائل التواصل على صورة الجسم في العالم العربي",
      "ما كشفته الدراسات العربية عن صورة الجسم — والأرقام مقلقة.",
      "2100-word article: Saudi Arabia body-image research, Egypt comparison, Arab cultural specificities (modesty paradox, wedding culture, family dynamics), whitening-product market as indicator. Closes with dokanelbanat.com Week 3 framework intro.",
      "Hero: illustrated Arab world map with embedded data-point annotations. Female silhouettes at key markers. Research-journal aesthetic with warm green and pink gradient overlay.",
      "campaign/blog/article-05-saudi-body-image-study.md"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 3 — SHIFT  |  Days 15-21  |  May 15-21, 2026
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 15 ──────────────────────────────────────────────────────────────────
    (15,"2026-05-15","Week 3 – Shift","Alternative Mindset",
      "Instagram","Carousel",
      "بداية التحول: اخترنا الوعي",
      "الوعي مش قرار لحظة — هو عادة يومية. وبتبدأ هنا.",
      "6-slide pivot carousel. Visual tone shifts to lighter (#fff5fb). Slide 1: 'أسبوع التحول' title with sunrise gradient. Slide 2: what the shift means (awareness not restriction). Slides 3-5: three mindset reframes. Slide 6: introduces 4 Pillars framework preview.",
      "Sunrise aesthetic: Arab woman stretching outdoors in morning light. Brand colour sunrise gradient (#ffbe3d to #ff6b4a to #ff009f). Clean, airy, hopeful. Rubik Bold Arabic motivational overlay.",
      "campaign/week-03-shift/posts/day-15-carousel.md"),

    (15,"2026-05-15","Week 3 – Shift","Alternative Mindset",
      "Facebook","Single Post",
      "التحول نحو الاستهلاك الواعي: من أين تبدئين؟",
      "إزاي تبدأي رحلة الوعي من غير ما تحسي بذنب على الماضي.",
      "Single warm-tone Facebook post. Long-form caption (180 words): non-judgmental reflection on past unconscious spending, reframe as learning, 3 concrete first steps for the conscious journey. Personal, empathetic brand voice.",
      "Warm flat-lay: journal open to blank page, 3 minimal skincare products, morning coffee, soft natural side-lighting. Cream and blush tones. Aspirational-real.",
      "campaign/week-03-shift/posts/day-15-single.md"),

    # DAY 16 ──────────────────────────────────────────────────────────────────
    (16,"2026-05-16","Week 3 – Shift","Alternative Mindset",
      "Instagram","Story Sequence",
      "فلسفة المينيماليزم في الجمال: أقل يساوي أكثر",
      "أقل = أكتر. فلسفة المينيماليزم في الجمال والحياة اليومية.",
      "5-story vertical sequence. Story 1: define beauty minimalism vs. deprivation. Story 2: capsule skincare (5 products). Story 3: capsule wardrobe (10 pieces). Story 4: 5-question routine audit. Story 5: CTA to dokanelbanat.com minimalism guide.",
      "Minimal aesthetic: white and cream full-bleed. Simple product groupings flat-lay. 'less is more' typography treatment. Clean geometric framing. Each story one concise message.",
      "campaign/week-03-shift/stories/day-16-story-sequence.md"),

    (16,"2026-05-16","Week 3 – Shift","Alternative Mindset",
      "Blog","Blog Article",
      "المينيماليزم والجمال: كيف تبنين روتيناً واعياً بخمسة منتجات فقط",
      "الروتين الواعي مش محتاج 20 منتج — 5 كفاية.",
      "1900-word minimalist beauty guide: capsule skincare philosophy, ingredient awareness basics, multi-use products, slow beauty movement, Arab skin types mini-guide, dokanelbanat.com curated 5-product starter list.",
      "Hero: flat-lay of exactly 5 skincare products on white marble, soft directional natural light. Cream and pink tones. Minimalist editorial photography. Intentional negative space.",
      "campaign/blog/article-06-minimalism-beauty.md"),

    # DAY 17 ──────────────────────────────────────────────────────────────────
    (17,"2026-05-17","Week 3 – Shift","Alternative Mindset",
      "Instagram","Carousel",
      "الروتين اليومي الواعي: خطوة بخطوة",
      "يومك من غير ما السوشيال ميديا تحكمه — روتين يومي واعي.",
      "7-slide practical carousel. Each slide = time block: Morning Ritual, Mid-Morning Check-in, Lunch Mindfulness, Afternoon Recharge, Evening Wind-down, Night Reflection, Weekly Review. Each has awareness check, intentional choice, gratitude micro-practice.",
      "Day-planner illustration: time blocks as clean calendar rows with warm icons. #ffbe3d morning, #3db5aa afternoon, #ff009f evening. Planner-book page texture background.",
      "campaign/week-03-shift/posts/day-17-carousel.md"),

    (17,"2026-05-17","Week 3 – Shift","Alternative Mindset",
      "Facebook","Story Sequence",
      "أول 30 دقيقة في يومك: لماذا هي الأهم؟",
      "أول 30 دقيقة في يومك — اللي بتقرري فيها مزاجك كله.",
      "4-story practical guide. Story 1: research-backed reason why first 30 minutes shape the day. Stories 2-3: phone-free morning steps with alternative habit suggestions. Story 4: 7-day morning challenge invitation with dokanelbanat.com community link.",
      "Morning photography stories: sunrise light through curtain, open journal, glass of water, no phone present. Warm cream and gold tones. Calm, inviting, peaceful.",
      "campaign/week-03-shift/stories/day-17-story-sequence.md"),

    # DAY 18 ──────────────────────────────────────────────────────────────────
    (18,"2026-05-18","Week 3 – Shift","Alternative Mindset",
      "Instagram","Thread",
      "الاكتفاء الذاتي: المرأة القوية التي تكتفي بنفسها",
      "الاكتفاء مش عزلة — هو قوة. الفرق بين الاثنين.",
      "8-post thread: define self-sufficiency vs. isolation, financial self-sufficiency basics, emotional self-sufficiency (no external validation), knowledge self-sufficiency, skill-building, community vs. dependency, Arab cultural context, dokanelbanat.com as resource hub.",
      "Thread card series: bold Arabic keyword centred per card (الاكتفاء / القوة / المعرفة etc.), supporting stat/insight below, #ff009f and #3db5aa alternating accent lines. Strong clean identity.",
      "campaign/week-03-shift/posts/day-18-thread.md"),

    (18,"2026-05-18","Week 3 – Shift","Alternative Mindset",
      "Instagram","Infographic",
      "أربعة أركان المرأة الواعية: النموذج الكامل",
      "4 أركان — 1 هوية. اكتشفي النموذج اللي بيبنيكي.",
      "Square 4-quadrant infographic (1080x1080). TL: Self-Sufficiency الاكتفاء الذاتي (#ff009f). TR: Conscious Awareness الوعي الاستهلاكي (#ff6b4a). BL: Intentional Routine الروتين المتعمد (#ffbe3d). BR: Empowering Community المجتمع الداعم (#3db5aa). Each: pillar icon + Arabic name + 3-word description.",
      "Four-quadrant geometric design: each section a distinct brand colour, centred icon and Arabic label. Clean sans-serif type. Symmetrical, balanced, memorable framework. White dividers.",
      "campaign/week-03-shift/infographics/day-18-infographic.md"),

    # DAY 19 ──────────────────────────────────────────────────────────────────
    (19,"2026-05-19","Week 3 – Shift","Alternative Mindset",
      "Instagram","Carousel",
      "فلسفة الجمال النظيف: Clean Beauty بالعربي",
      "Clean Beauty مش ترند — هو حق. المرأة العربية والمكونات النظيفة.",
      "6-slide carousel. Slide 1: clean beauty definition in Arab context. Slides 2-3: 6 harmful ingredients (Arabic INCI names) with red-flag icons + safe alternatives. Slide 4: Arab-specific skin concerns and clean solutions. Slide 5: how to read Arabic product label. Slide 6: dokanelbanat.com clean picks.",
      "Clean beauty flat-lay: white marble, fresh botanicals (aloe, rose petals, argan), Arabic ingredient label visible. #3db5aa for 'clean', red for 'avoid'. Scientific-beautiful aesthetic.",
      "campaign/week-03-shift/posts/day-19-carousel.md"),

    (19,"2026-05-19","Week 3 – Shift","Alternative Mindset",
      "Facebook","Thread",
      "كيف تقرئين مكونات أي منتج تجميل بالعربي؟",
      "قبل ما تشتري أي منتج تجميل — اقرأي ده الأول.",
      "Facebook thread: step-by-step ingredient reading guide. Root post explains why it matters. 6 comments: INCI name basics, order of concentration, 3 always-avoid ingredients, 3 look-for ingredients, Arabic vs INCI name reference table, dokanelbanat.com ingredient checker tool link.",
      "Thread visual cards: product ingredient list mockup with highlighted warning ingredients in red, green-tick safe alternatives. Magnifying glass illustration. Clean educational flat design.",
      "campaign/week-03-shift/posts/day-19-thread.md"),

    # DAY 20 ──────────────────────────────────────────────────────────────────
    (20,"2026-05-20","Week 3 – Shift","Alternative Mindset",
      "Instagram","Story Sequence",
      "معايير الجمال الحقيقية: من قال إنها صح؟",
      "من قرر إن الشعر الناعم أحسن؟ ومن قرر إن البشرة الفاتحة أجمل؟",
      "5-story deconstruction sequence. Story 1: 'من قرر؟' powerful opening. Story 2: colonial history of Arab beauty standards (brief). Story 3: diversity showcase — 4 types of natural Arab beauty. Story 4: media's standard-manufacturing role. Story 5: empowerment reframe + question sticker.",
      "Diverse Arab women portrait series (illustrated): each portrait different natural features. Each has a deconstructed beauty standard as crossed-out overlay text. Vibrant colours per portrait. Celebratory.",
      "campaign/week-03-shift/stories/day-20-story-sequence.md"),

    (20,"2026-05-20","Week 3 – Shift","Alternative Mindset",
      "Blog","Blog Article",
      "معايير الجمال العربية: من صنعها ولماذا؟",
      "الجمال العربي تحت المجهر — من وضع المعايير ولمصلحة من؟",
      "2000-word article: colonial-era imposition of Eurocentric standards, media's manufacturing role, skin-lightening industry ($8.6B globally), natural Arab beauty diversity by geography, framework for reclaiming standards.",
      "Hero: illustrated collage of diverse Arab women in natural everyday settings (not fashion). Warm earth tones, deconstructed beauty-product imagery in background. Empowering editorial quality.",
      "campaign/blog/article-07-arab-beauty-standards.md"),

    # DAY 21 ──────────────────────────────────────────────────────────────────
    (21,"2026-05-21","Week 3 – Shift","Alternative Mindset",
      "Instagram","Carousel",
      "ملخص الأسبوع الثالث: التحول اكتمل",
      "3 أسابيع — والتحول بدأ. إيه اللي تغير في طريقة تفكيرك؟",
      "6-slide recap. Lighter colour tone. Slide 1: 'Week 3 Shift Complete' with airy gradient. Slides 2-5: one insight per shift topic (minimalism, 4 pillars, clean beauty, real standards). Slide 6: Week 4 Empower teaser 'أنتِ مكتملة' + CTA.",
      "Week 3 content mosaic thumbnails in lighter colour treatment. Airy open aesthetic. Bold Week 4 teaser slide with energising typography as final card.",
      "campaign/week-03-shift/posts/day-21-carousel.md"),

    (21,"2026-05-21","Week 3 – Shift","Alternative Mindset",
      "Facebook","Infographic",
      "قبل الوعي وبعده: جدول المقارنة الكامل",
      "حياتك قبل الوعي الاستهلاكي وبعده — الفرق واضح.",
      "Before/After comparison infographic (Facebook landscape). Two columns: قبل الوعي (dark, cluttered) vs بعد الوعي (clean, calm). 8 comparison rows: spending, social media time, self-image, buying triggers, community quality, knowledge, daily routine, brand loyalty. Icon per row.",
      "Split-table design: left half dark/chaotic, right half clean/calm. Same Arab woman illustrated in both halves in contrasting energy states. Brand colour transformation story visible.",
      "campaign/week-03-shift/infographics/day-21-infographic.md"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 4 — EMPOWER  |  Days 22-30  |  May 22-30, 2026
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 22 ──────────────────────────────────────────────────────────────────
    (22,"2026-05-22","Week 4 – Empower","You Are Complete",
      "Instagram","Carousel",
      "أنتِ مكتملة: البيان الأول",
      "مش محتاجة أي منتج علشان تبقي مكتملة. أنتِ مكتملة دلوقتي.",
      "7-slide manifesto carousel. Maximum typographic impact — no product imagery. Slide 1: 'أنتِ مكتملة' full-bleed bold on #ff009f. Slides 2-6: each completeness dimension (inner strength, knowledge, authentic beauty, community, purpose). Slide 7: dokanelbanat.com as home.",
      "Power-poster typographic design: full-frame Arabic bold text, each slide a different brand gradient. No photographs — pure type and colour. Manifesto energy.",
      "campaign/week-04-empower/posts/day-22-carousel.md"),

    (22,"2026-05-22","Week 4 – Empower","You Are Complete",
      "Facebook","Single Post",
      "رسالة dokanelbanat.com: لماذا أنشأنا هذه المنصة",
      "السبب الحقيقي وراء dokanelbanat.com — مش متجر. مجتمع.",
      "Founder's-voice Facebook post. Long-form narrative (250 words): origin story of dokanelbanat.com, problem observed, gap in market, community vision, what makes it different. Personal, authentic, mission-led. Invites community to share their 'why'.",
      "Warm editorial portrait: woman at desk with open journal and laptop, dokanelbanat.com subtle on screen. Candid natural light, #fff5fb tones. Real not staged.",
      "campaign/week-04-empower/posts/day-22-single.md"),

    # DAY 23 ──────────────────────────────────────────────────────────────────
    (23,"2026-05-23","Week 4 – Empower","You Are Complete",
      "Instagram","Story Sequence",
      "قصص من مجتمعنا: نساء اخترن الوعي",
      "مش بنحكي عنهم — بنحكي معاهم. قصص حقيقية من مجتمعنا.",
      "5-story community spotlight. Each story = one member (anonymised or consented): challenge in 1 line, shift moment, outcome. Story 5: submissions call 'شاركي قصتك مع dokanelbanat.com'.",
      "Portrait spotlight: soft-vignette illustrated portrait, first-person Arabic quote pull, #ff009f accent frame. Community member name/alias in Rubik Bold. Warm, human.",
      "campaign/week-04-empower/stories/day-23-story-sequence.md"),

    (23,"2026-05-23","Week 4 – Empower","You Are Complete",
      "Blog","Blog Article",
      "نساء رائدات: قصص من مجتمع dokanelbanat.com",
      "نساء عربيات اخترن الوعي — قصصهم بتغير مجتمعات.",
      "1800-word community spotlight article. 4-5 stories of Arab women entrepreneurs and conscious consumers. Each: background, challenge, mindset-shift catalyst, measurable change. dokanelbanat.com as community catalyst.",
      "Hero: illustrated group portrait of 4-5 diverse Arab women in professional natural settings. Warm community energy, not corporate. Editorial quality with brand colour accents.",
      "campaign/blog/article-08-community-women-stories.md"),

    # DAY 24 ──────────────────────────────────────────────────────────────────
    (24,"2026-05-24","Week 4 – Empower","You Are Complete",
      "Instagram","Infographic",
      "النساء رائدات الأعمال في العالم العربي: أرقام الإلهام",
      "المرأة العربية رائدة أعمال — والأرقام بتقول حاجة كبيرة.",
      "Portrait infographic on Arab women entrepreneurs: % of MENA startups founded by women, Egypt female entrepreneurship growth rate, top sectors (e-commerce, beauty, education, food), key barriers, dokanelbanat.com as support ecosystem.",
      "Bold entrepreneurship infographic: rising arrow charts, female silhouette with business elements (laptop, storefront, product). #ff009f and #ffbe3d dominant. Arabic stat callouts. Confident aspirational.",
      "campaign/week-04-empower/infographics/day-24-infographic.md"),

    (24,"2026-05-24","Week 4 – Empower","You Are Complete",
      "Facebook","Carousel",
      "كيف تبدئين مشروعك الصغير من الصفر؟",
      "من الفكرة للبيعة الأولى — دليل عملي للمرأة العربية.",
      "6-slide practical startup carousel. Steps: idea validation (3 questions), minimal viable product, first sales channel options, pricing psychology basics, building first audience, how dokanelbanat.com fits every step.",
      "Step-by-step guide cards: steps 1-6, icon illustration per step, #ff6b4a orange dominant. Professional but approachable. Small business imagery (handmade product, phone camera, simple storefront).",
      "campaign/week-04-empower/posts/day-24-carousel.md"),

    # DAY 25 ──────────────────────────────────────────────────────────────────
    (25,"2026-05-25","Week 4 – Empower","You Are Complete",
      "Instagram","Thread",
      "منظومة dokanelbanat.com: كل ما تحتاجينه في مكان واحد",
      "واحدة سألتني: إيه اللي عنده dokanelbanat.com فعلاً؟ — اتفضلي.",
      "8-post thread. Each post = one ecosystem pillar: awareness content library, educational resources, community forum, curated conscious product discovery, entrepreneur support programme, expert network, events & workshops, weekly newsletter. Pillar name + what you get + how to access.",
      "Ecosystem map thread cards: hub-and-spoke diagram with dokanelbanat.com logo at centre. Each pillar as labelled spoke. Consistent card design. Brand colour per pillar. Final card is full map.",
      "campaign/week-04-empower/posts/day-25-thread.md"),

    (25,"2026-05-25","Week 4 – Empower","You Are Complete",
      "Blog","Blog Article",
      "منظومة dokanelbanat.com: كيف تستفيدين من كل خدمة؟",
      "كل ما تقدمه dokanelbanat.com — وكيف تستخدمينه لصالحك.",
      "2000-word complete ecosystem guide. Each section = one platform offering with specific use-cases, how to access, real benefit for Arab women. Written as practical onboarding guide for new members. SEO-optimised for brand keywords.",
      "Hero: dokanelbanat.com interface mockup surrounded by community icons, educational elements, curated product imagery. Brand-accurate colour palette. Editorial software-screenshot aesthetic.",
      "campaign/blog/article-09-dokanelbanat-ecosystem.md"),

    # DAY 26 ──────────────────────────────────────────────────────────────────
    (26,"2026-05-26","Week 4 – Empower","You Are Complete",
      "Instagram","Carousel",
      "بناء مجتمع مخلص: الفرق بين الجمهور والمجتمع",
      "مليون فولوور مش مجتمع — وألف امرأة واعية أقوى منهم.",
      "7-slide community-building carousel. Slide 1: audience vs community visual contrast. Slides 2-5: 4 pillars of real community (shared values, two-way dialogue, mutual support, collective growth). Slide 6: dokanelbanat.com community principles. Slide 7: join CTA + first 7 days as member.",
      "Community visualisation: concentric-circle diagram (depth vs breadth). Warm illustrated human figures connected by lines. #3db5aa green dominant. Conversations and connections depicted.",
      "campaign/week-04-empower/posts/day-26-carousel.md"),

    (26,"2026-05-26","Week 4 – Empower","You Are Complete",
      "Facebook","Story Sequence",
      "كيف تنضمين لمجتمع dokanelbanat.com؟",
      "3 خطوات بس — وتبقي جزء من أقوى مجتمع نساء عربيات.",
      "3-story onboarding sequence. Story 1: member benefits visual 'ما الذي ستحصلين عليه'. Story 2: step-by-step joining guide with screen mockup. Story 3: your first community action 'عرّفي عن نفسك' with dokanelbanat.com link.",
      "Onboarding welcome design: open-door visual metaphor with warm glow inside. Step icons clean minimal. Community member avatar cluster. Welcoming brand colour gradient.",
      "campaign/week-04-empower/stories/day-26-story-sequence.md"),

    # DAY 27 ──────────────────────────────────────────────────────────────────
    (27,"2026-05-27","Week 4 – Empower","You Are Complete",
      "Instagram","Story Sequence",
      "الاستهلاك الواعي في 30 يوم: رحلتنا معاً",
      "30 يوم من الوعي — والرحلة مش خلصت، هي بس بدأت.",
      "5-story campaign retrospective. Story 1: campaign arc overview (4 weeks visual map). Story 2: most impactful Week 1 reveal + community reaction. Story 3: most shared Week 2 stat. Story 4: most resonant Week 3 shift moment. Story 5: Week 4 highlight + question sticker.",
      "Film-strip retrospective: each story shows mini-thumbnails of campaign highlights in that phase. Nostalgic-meets-forward aesthetic. Brand gradient evolution dark (Week 1) to light (Week 4).",
      "campaign/week-04-empower/stories/day-27-story-sequence.md"),

    (27,"2026-05-27","Week 4 – Empower","You Are Complete",
      "Facebook","Thread",
      "30 يوم من الوعي: 10 دروس كبيرة تعلمناها معاً",
      "كل يوم علمنا حاجة — دي أهم 10 دروس من 30 يوم.",
      "Facebook thread: 10 campaign lessons. Root post frames retrospective. Each of 10 comments = one lesson: lesson statement, the Day it came from, community engagement data point validating it, forward implication for daily life.",
      "Campaign learning cards: numbered lessons in clean list format. #ff009f bullet circles. High-contrast shareable-screenshot design. Summary of 30-day journey.",
      "campaign/week-04-empower/posts/day-27-thread.md"),

    # DAY 28 ──────────────────────────────────────────────────────────────────
    (28,"2026-05-28","Week 4 – Empower","You Are Complete",
      "Instagram","Carousel",
      "بناء البراند الشخصي للمرأة العربية",
      "براندك الشخصي مش لوجو ولا كولر — هو قيمتك في السوق.",
      "6-slide personal branding carousel. Slide 1: redefine personal brand for Arab women. Slide 2: values-based positioning (not just aesthetics). Slide 3: authentic storytelling framework. Slide 4: LinkedIn for Arab professional women quick guide. Slide 5: consistency over virality. Slide 6: dokanelbanat.com as brand amplifier.",
      "Personal brand builder carousel: clean professional design, woman-at-work illustrated, brand identity elements depicted (colour swatches, font samples, content calendar). #ff009f accent. Empowering business aesthetic.",
      "campaign/week-04-empower/posts/day-28-carousel.md"),

    (28,"2026-05-28","Week 4 – Empower","You Are Complete",
      "Blog","Blog Article",
      "البراند الشخصي للمرأة العربية: دليل البداية في 2026",
      "كيف تبنين براند شخصي أصيل كامرأة عربية في 2026؟",
      "2100-word personal branding guide: discovering unique positioning, values audit exercise, authentic online presence on LinkedIn and Instagram, Arab cultural context in professional branding, community leverage strategy, 90-day action plan.",
      "Hero: confident Arab professional woman at whiteboard with brand strategy elements sketched behind. Warm office setting. Brand colours in environment design. Inspirational, real.",
      "campaign/blog/article-10-personal-brand-arab-women.md"),

    # DAY 29 ──────────────────────────────────────────────────────────────────
    (29,"2026-05-29","Week 4 – Empower","You Are Complete",
      "Instagram","Thread",
      "دليل الاستهلاك الواعي الذي يبقى معكِ للأبد",
      "مش حملة — طريقة حياة. الدليل الكامل للاستهلاك الواعي.",
      "8-post evergreen thread: monthly spending audit protocol, social media feed curation guide, product ingredient checklist, community validation before purchase, supporting women-owned brands, daily gratitude-body practice, quarterly mindset review, annual brand-and-media detox guide.",
      "Evergreen handbook thread cards: each card styled as a 'chapter' in a clean handbook. Chapter number bold, #3db5aa green for permanent feel. Practical icons, Arabic chapter titles. Could be printable cards.",
      "campaign/week-04-empower/posts/day-29-thread.md"),

    (29,"2026-05-29","Week 4 – Empower","You Are Complete",
      "Facebook","Infographic",
      "خريطة قرار الشراء الواعي: 7 خطوات قبل أي منتج",
      "كل قرار شراء بيمر بـ7 خطوات — هل بتمشي عليهم؟",
      "Tall decision-flowchart infographic. 7-step purchase decision tree: (1) هل أحتاجه فعلاً؟ (2) بحث المكونات (3) مراجعة المجتمع (4) توافق القيم (5) الميزانية (6) التأثير البيئي (7) اشتري أو انتظري. Yes/No branches. Shareable poster format.",
      "Flowchart poster: clean Arabic decision tree. Yes branches in #3db5aa green, pause branches in #ff009f, research branches in #ffbe3d. White background, clear typography, infographic-quality design.",
      "campaign/week-04-empower/infographics/day-29-infographic.md"),

    # DAY 30 ──────────────────────────────────────────────────────────────────
    (30,"2026-05-30","Week 4 – Empower","You Are Complete",
      "Instagram","Carousel",
      "يوم 30: أنتِ مكتملة — البيان النهائي",
      "30 يوم، ومجتمع كامل. أنتِ مكتملة وجاهزة تغيري الوعي.",
      "8-slide campaign finale. Slide 1: bold finale statement on #ff009f. Slides 2-4: 3-phase journey recap (Expose/Educate/Shift). Slide 5: community achievement metrics. Slide 6: 4 Pillars visual (final version). Slide 7: what comes next. Slide 8: closing manifesto 'أنتِ مكتملة. دلوقتي وللأبد.'",
      "Grand finale carousel: campaign's strongest visual execution. Full-bleed brand gradients, #ffbe3d gold celebration accents, community portrait montage, closing manifesto in maximum typographic weight. Cinematic.",
      "campaign/week-04-empower/posts/day-30-carousel.md"),

    (30,"2026-05-30","Week 4 – Empower","You Are Complete",
      "Blog","Blog Article",
      "30 يوم من الوعي: الرسالة الختامية لمجتمع dokanelbanat.com",
      "الرسالة الأخيرة — وبداية كل شيء.",
      "2500-word campaign closing manifesto: (1) what we exposed Week 1, (2) what data confirmed Week 2, (3) mindset shift built Week 3, (4) community we became Week 4. Closes with dokanelbanat.com long-term commitments. Manifesto-quality writing, designed to be shared.",
      "Hero: wide community illustration — circle of diverse Arab women, sunrise behind, empowering atmosphere. All brand colours harmonised. Cinematic campaign-finale quality.",
      "campaign/blog/article-11-campaign-closing-manifesto.md"),

    (30,"2026-05-30","Week 4 – Empower","You Are Complete",
      "Facebook","Story Sequence",
      "شكراً — وما الذي يأتي بعد ذلك مع dokanelbanat.com",
      "الحملة خلصت — المجتمع لأ. إيه اللي بيجي بعد كده؟",
      "4-story closing sequence. Story 1: thank-you with campaign stats (reach/engagement/community growth). Story 2: what comes next from dokanelbanat.com. Story 3: how to stay connected (newsletter, community, social). Story 4: 'شاركي تحولك' CTA with dokanelbanat.com hashtag.",
      "Closing celebration stories: confetti-particle effect in brand colours. Warm gratitude aesthetic. Forward-looking imagery (horizon, open door). Newsletter CTA card as final frame — clean and compelling.",
      "campaign/week-04-empower/stories/day-30-story-sequence.md"),
]

# ── Build Workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "30-Day Calendar"

# ── Styles ────────────────────────────────────────────────────────────────────
hdr_fill  = PatternFill("solid", fgColor=C_PRIMARY)
hdr_font  = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
alt_fill  = PatternFill("solid", fgColor=C_ALT)
plain_fill= PatternFill("solid", fgColor=C_WHITE)
ctr_al    = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft_al    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin      = Side(style="thin", color="DDDDDD")
bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

week_fills = {
    "Week 1 – Bold":    PatternFill("solid", fgColor="FFE5F5"),
    "Week 2 – Educate": PatternFill("solid", fgColor="E5F7FF"),
    "Week 3 – Shift":   PatternFill("solid", fgColor="E5FFE8"),
    "Week 4 – Empower": PatternFill("solid", fgColor="FFF8E5"),
}
platform_clrs = {"Instagram":"C13584","Facebook":"1877F2","Blog":"FF6B4A"}
ctype_clrs    = {
    "Carousel":"FF009F","Single Post":"333333","Story Sequence":"FF2D55",
    "Infographic":"3DB5AA","Thread":"FF6B4A","Blog Article":"996600"
}

# ── Header Row ────────────────────────────────────────────────────────────────
ws.append(HEADERS)
ws.row_dimensions[1].height = 32
for ci in range(1, len(HEADERS)+1):
    c = ws.cell(row=1, column=ci)
    c.fill=hdr_fill; c.font=hdr_font; c.alignment=ctr_al; c.border=bdr

# ── Data Rows ─────────────────────────────────────────────────────────────────
for ri, rd in enumerate(ROWS, start=2):
    ws.append(list(rd) + ["Ready to Publish"])
    ws.row_dimensions[ri].height = 58

    wk      = rd[2]
    is_alt  = (ri % 2 == 0)
    rf      = alt_fill if is_alt else week_fills.get(wk, plain_fill)
    plat    = rd[4]
    ctype   = rd[5]

    for ci in range(1, len(HEADERS)+1):
        c = ws.cell(row=ri, column=ci)
        c.fill   = rf
        c.border = bdr
        c.alignment = ctr_al if ci in (1,2,3,5,6,12) else lft_al

        if ci == 1:
            c.font = Font(name="Calibri", bold=True, size=12, color=C_PRIMARY)
        elif ci == 5:
            c.font = Font(name="Calibri", bold=True, size=10, color=platform_clrs.get(plat,"222222"))
        elif ci == 6:
            c.font = Font(name="Calibri", bold=True, size=10, color=ctype_clrs.get(ctype,"222222"))
        elif ci == 12:
            c.font = Font(name="Calibri", bold=True, size=10, color="3DB5AA")
        else:
            c.font = Font(name="Calibri", size=10)

ws.freeze_panes = "A2"

# ── Column Widths ─────────────────────────────────────────────────────────────
WIDTHS={1:6,2:14,3:17,4:22,5:13,6:17,7:38,8:45,9:62,10:58,11:52,12:18}
for ci in range(1, len(HEADERS)+1):
    ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(ci,20)

# ── Summary Sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Campaign Summary")
tc = {}; wc = {}; pc = {}
for r in ROWS:
    tc[r[5]]=tc.get(r[5],0)+1
    wc[r[2]]=wc.get(r[2],0)+1
    pc[r[4]]=pc.get(r[4],0)+1

summary=[
    ["dokanelbanat.com — 30-Day Campaign Calendar",""],["",""],
    ["Metric","Value"],
    ["Total Content Rows", len(ROWS)],["Campaign Days",30],
    ["Start Date","2026-05-01"],["End Date","2026-05-30"],
    ["Platforms",", ".join(sorted(pc.keys()))],
    ["All Status","Ready to Publish"],["",""],
    ["CONTENT TYPE","COUNT"],
]
for k,v in sorted(tc.items()): summary.append([k,v])
summary+=[["",""],["WEEK","ROWS"]]
for k,v in sorted(wc.items()): summary.append([k,v])
summary+=[["",""],["PLATFORM","ROWS"]]
for k,v in sorted(pc.items()): summary.append([k,v])

for sr in summary:
    ws2.append(sr)
for ri2,sr in enumerate(summary,start=1):
    for ci2 in range(1,3):
        c=ws2.cell(row=ri2,column=ci2)
        c.border=bdr; c.alignment=lft_al
        if sr and len(sr)>0 and sr[0] in (
            "dokanelbanat.com — 30-Day Campaign Calendar",
            "Metric","CONTENT TYPE","WEEK","PLATFORM"
        ):
            c.fill=hdr_fill; c.font=hdr_font
        else:
            c.font=Font(name="Calibri",size=10)
ws2.column_dimensions["A"].width=35
ws2.column_dimensions["B"].width=20
ws2.row_dimensions[1].height=28

# ── Save & Verify ─────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)

p=pathlib.Path(OUTPUT_PATH)
if p.exists():
    sz=p.stat().st_size
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Size:  {sz:,} bytes ({sz/1024:.1f} KB)")
    print(f"Rows:  {len(ROWS)}")
    print(f"Types: {tc}")
    print(f"Weeks: {wc}")
    print(f"Platforms: {pc}")
    print("\nCALENDAR DONE")
else:
    print("ERROR: File not saved!")
    import sys; sys.exit(1)

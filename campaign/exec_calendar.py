#!/usr/bin/env python3
"""
dokanelbanat.com — 30-Day Content Calendar Builder
Outputs: D:\claude-Projects\dokanelbanat\campaign\calendar.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys

OUTPUT_PATH = r"D:\claude-Projects\dokanelbanat\campaign\calendar.xlsx"
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# ── Brand Palette ─────────────────────────────────────────────────────────────
C_PRIMARY   = "FF009F"
C_SECONDARY = "FF2D55"
C_ORANGE    = "FF6B4A"
C_YELLOW    = "FFBE3D"
C_GREEN     = "3DB5AA"
C_LIGHT     = "FFF5FB"
C_WHITE     = "FFFFFF"
C_ALT       = "FFF0F8"   # alternating row tint

# ── Column Headers ────────────────────────────────────────────────────────────
HEADERS = [
    "Day", "Date", "Week", "Theme",
    "Platform", "Content Type", "Topic (Arabic)",
    "Hook Line", "Design Brief Summary",
    "Image Prompt Summary", "File Path", "Status"
]

# ── Full Calendar Data ────────────────────────────────────────────────────────
# Tuple order matches HEADERS (minus Status, appended automatically):
# (Day, Date, Week, Theme, Platform, ContentType, TopicArabic,
#  HookLine, DesignBrief, ImagePrompt, FilePath)

ROWS = [

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 1 — BOLD & PROVOCATIVE  |  Days 1–7  |  May 1–7, 2026
    # Topics: consumer machine, Instagram & body image, $677B beauty industry,
    #         housewife advertising history, influencer purchase psychology,
    #         social comparison theory, Arab women consumer spending
    # ══════════════════════════════════════════════════════════════════════════

    # ── DAY 1 ─────────────────────────────────────────────────────────────────
    (1, "2026-05-01", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Carousel",
     "ماكينة الإنفاق العالمية على المرأة",
     "كل ما تحسي إنك ناقصة — في حد بيكسب منها.",
     "6-slide carousel. Slide 1: bold #ff009f headline on near-black background. "
     "Slides 2–5: one data stat per slide — large numeral in #ffbe3d, icon below, "
     "short explanatory Arabic text. Slide 6: dokanelbanat.com URL + follow CTA. "
     "Rubik Bold for all headlines, MarkaziText for sub-copy.",
     "Giant chrome industrial machine with conveyor belt; shopping bags and beauty "
     "products flow through it and land on a silhouetted woman. Neon pink (#ff009f) "
     "lighting on deep black background. Editorial, high-contrast.",
     "campaign/week-01-bold/posts/day-01-carousel.md"),

    (1, "2026-05-01", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Single Post",
     "لماذا تشعرين دائماً بالنقص؟",
     "السؤال اللي محدش بيسأله — ليه إحساس النقص ده مش بيروح أبداً؟",
     "Single square image. Left half: chaotic consumerist collage (pink neon). "
     "Right half: calm Arab woman looking away. Bold Arabic question runs across "
     "both halves. #ff009f accent strip down center. Facebook caption 150 words.",
     "Split-frame photograph: left = chaotic pink-neon shopping mall interior; "
     "right = serene Arab woman in natural light looking to horizon. Sharp split line "
     "with brand pink as divider. Cinematic ratio.",
     "campaign/week-01-bold/posts/day-01-single.md"),

    # ── DAY 2 ─────────────────────────────────────────────────────────────────
    (2, "2026-05-02", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Story Sequence",
     "انستجرام وصورة الجسم عند المرأة العربية",
     "دقيقتين على انستجرام = تغيير في صورتك عن نفسك.",
     "5-story vertical sequence (1080×1920). Story 1: shocking opening stat on "
     "#ff009f full-bleed background. Stories 2–3: scrolling stat cards with "
     "progress bar. Story 4: 'هل حصل معاكي؟' poll sticker. Story 5: swipe-up / "
     "link-sticker to blog article. Gradient transitions between stories.",
     "Phone screen mockup filled with filtered body images from Instagram; "
     "a spreading crack-glitch effect radiates from the screen. Deep pink and "
     "charcoal tones. Close-up perspective.",
     "campaign/week-01-bold/stories/day-02-story-sequence.md"),

    (2, "2026-05-02", "Week 1 – Bold", "Expose the Machine",
     "Blog", "Blog Article",
     "انستجرام وأزمة صورة الجسم عند المرأة العربية",
     "كيف تحوّل انستجرام المرأة العربية إلى زبونة دائمة؟",
     "1 800-word deep-dive. Sections: (1) research overview, (2) psychological "
     "triggers activated by filtered content, (3) Arab-specific data and cultural "
     "context, (4) how the algorithm exploits body insecurity, (5) what "
     "dokanelbanat.com offers as alternative. Internal links + CTA at end.",
     "Hero illustration: Arab woman holding a phone; her mirror reflection shows "
     "an idealised filtered body. Painterly style. #ff009f dominant with cream "
     "accents. Wide-format blog hero.",
     "campaign/blog/article-01-instagram-body-image.md"),

    # ── DAY 3 ─────────────────────────────────────────────────────────────────
    (3, "2026-05-03", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Infographic",
     "صناعة الجمال: 677 مليار دولار — وكيف بُنيت على النقص",
     "$677,000,000,000 — ده حجم الصناعة اللي بتتبنى على إحساسك بالنقص.",
     "Tall portrait infographic (1080×1920). Decade-by-decade timeline from "
     "1920s to 2026: industry revenue figure per era + key manipulation tactic "
     "of that decade. Bold numerals in #ffbe3d on dark charcoal. dokanelbanat.com "
     "watermark footer. Rubik Bold labels.",
     "Vertical timeline poster: each era has a miniature vintage-to-modern ad "
     "style and a glowing golden revenue figure. Dark charcoal background, pink "
     "timeline spine, decade labels in white.",
     "campaign/week-01-bold/infographics/day-03-infographic.md"),

    (3, "2026-05-03", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Story Sequence",
     "تاريخ إعلانات المرأة ربة المنزل",
     "من 'ابشري يا ست البيت' لـ'بشرتك مش كفاية' — الإعلان دايماً عارف يضغط على الجرح.",
     "4-story Facebook stories sequence. Story 1: vintage housewife ad aesthetic "
     "with sepia overlay. Stories 2–3: decade jump cards (50s → 80s → 2020s) "
     "showing evolved manipulation. Story 4: poll — 'شايفة نفسك في الإعلانات؟'",
     "Collage of vintage Arabic newspaper housewife advertisements transitioning "
     "via film-burn effect to contemporary social media beauty ads. Sepia to neon "
     "pink color progression.",
     "campaign/week-01-bold/stories/day-03-story-sequence.md"),

    # ── DAY 4 ─────────────────────────────────────────────────────────────────
    (4, "2026-05-04", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Carousel",
     "سيكولوجية الشراء عند الإنفلوانسر",
     "مش بتشتري المنتج — بتشتري النسخة اللي إنتي عايزة تبقيها.",
     "7-slide carousel. Slide 1: influencer lifestyle still with overlay text. "
     "Slides 2–6: one psychological trigger per slide — FOMO, social proof, "
     "aspirational identity, scarcity urgency, parasocial bond. Each slide: "
     "trigger name bold in #ff009f, brief explanation, a real-world example. "
     "Slide 7: awareness CTA + dokanelbanat.com.",
     "Stylised illustration: Arab woman gazing into glowing phone; her reflection "
     "in the screen is a glamourised different version of herself, luxury-styled. "
     "Neon pink dominant, dark surround, dream-sequence atmosphere.",
     "campaign/week-01-bold/posts/day-04-carousel.md"),

    (4, "2026-05-04", "Week 1 – Bold", "Expose the Machine",
     "Blog", "Blog Article",
     "سيكولوجية الشراء: لماذا نصدق الإنفلوانسر؟",
     "العلم وراء زر 'اشتري دلوقتي' وليه دماغك مش قادرة تقاومه.",
     "2 000-word article. Covers: parasocial relationships theory, aspirational "
     "identity consumption, dopamine-reward loops in online shopping, FOMO "
     "mechanics, FTC disclosure failures in Arab markets, case studies with Arab "
     "influencer examples. Actionable awareness section at close.",
     "Hero: extreme close-up of an Arab woman's eye reflecting a glowing 'Buy Now' "
     "button. Cinematic lighting, shallow depth of field. Pink and gold tones, "
     "psychological tension.",
     "campaign/blog/article-02-influencer-psychology.md"),

    # ── DAY 5 ─────────────────────────────────────────────────────────────────
    (5, "2026-05-05", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Thread",
     "نظرية المقارنة الاجتماعية وتأثيرها على المرأة العربية",
     "ليه ما تقدريش تبطلي تقارني نفسك — وده مش عيب فيكي.",
     "8-post carousel-thread hybrid. Post 1: Festinger's 1954 theory intro. "
     "Post 2: upward vs downward comparison. Post 3: social media amplification "
     "mechanism. Post 4: body image link — research data. Post 5: self-esteem "
     "impact stats. Post 6: Arab cultural comparison layer (family + community). "
     "Post 7: breaking the cycle (awareness steps). Post 8: dokanelbanat.com "
     "community as counter-space. Consistent card design across all 8.",
     "Series of 8 minimal cards: Arabic key phrase centred on clean white/blush "
     "background. Subtle geometric pattern. Quote or stat in #ff009f beneath each "
     "phrase. Unified thread visual identity.",
     "campaign/week-01-bold/posts/day-05-thread.md"),

    (5, "2026-05-05", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Single Post",
     "المقارنة الاجتماعية في السياق العربي",
     "المجتمع بيعلمنا نقارن — والسوشيال ميديا بتتاجر فيها.",
     "Single educational Facebook post with statistics overlay on warm background. "
     "Long-form caption (200 words) exploring how Arab cultural norms of comparison "
     "(family gatherings, weddings) are turbo-charged by social platforms.",
     "Two Arab women side-by-side (illustrated): one showing phone to the other. "
     "The phone screen glows with comparison imagery. Warm orange (#ff6b4a) tones. "
     "Editorial flat illustration style.",
     "campaign/week-01-bold/posts/day-05-single.md"),

    # ── DAY 6 ─────────────────────────────────────────────────────────────────
    (6, "2026-05-06", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Story Sequence",
     "إنفاق المرأة العربية على منتجات التجميل",
     "إيه اللي بتنفقيه فعلاً — وعلى إيه بالظبط؟",
     "5-story sequence. Story 1: shocking monthly spend statistic full-bleed. "
     "Stories 2–3: category breakdown (skincare / makeup / hair / diet). Story 4: "
     "interactive quiz sticker 'كام صرفتي الشهر ده على التجميل؟'. Story 5: "
     "awareness reframe + dokanelbanat.com link.",
     "Animated-feel story frames: gold coins and beauty product icons floating on "
     "#ff009f background. Arabic numerals bold white. Calculator and sparkle emoji "
     "integrated. Clean, energetic composition.",
     "campaign/week-01-bold/stories/day-06-story-sequence.md"),

    (6, "2026-05-06", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Infographic",
     "أين تذهب أموال المرأة العربية؟",
     "خريطة الإنفاق — من جيبك لجيب مين؟",
     "Horizontal spending-breakdown infographic. Pie chart (centre) + category "
     "legend: skincare 35%, makeup 25%, hair 18%, diet & supplements 12%, fragrance "
     "10%. Each slice in a distinct brand-palette colour. Arabic category labels. "
     "Data source footnote. dokanelbanat.com footer.",
     "Overhead flat-lay arranged as a pie chart: beauty products physically grouped "
     "by category, each group on a colour-coded section. White marble base, top-down "
     "shot, natural lighting.",
     "campaign/week-01-bold/infographics/day-06-infographic.md"),

    # ── DAY 7 ─────────────────────────────────────────────────────────────────
    (7, "2026-05-07", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Carousel",
     "ملخص الأسبوع الأول: الماكينة انكشفت",
     "أسبوع كامل من الحقايق — إيه اللي صدمك أكتر؟",
     "6-slide week-recap carousel. Slide 1: bold 'Week 1 Recap' title card with "
     "#ff009f full-bleed. Slides 2–5: one key takeaway per topic day (machine, "
     "Instagram, influencer psychology, comparison, spending). Slide 6: teaser for "
     "Week 2 — 'الأسبوع الجاي: الأرقام والبيانات' + follow CTA.",
     "Mosaic grid of mini-thumbnail previews from Week 1 posts arranged in 2×3 "
     "layout. Bold WEEK 1 RECAP headline overlaid. #ff009f dominant. Rubik Bold "
     "typography throughout.",
     "campaign/week-01-bold/posts/day-07-carousel.md"),

    (7, "2026-05-07", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Thread",
     "أسئلة الأسبوع الأول من المجتمع",
     "إنتوا سألتوا — دي إجاباتنا. الحوار الحقيقي بيبدأ هنا.",
     "Facebook thread post. Root post asks community for their biggest Week 1 "
     "question. Top 5 pre-prepared answer comments pinned in order. Each comment "
     "cites a specific post or stat from the week. Engagement loop design.",
     "Community-conversation design: overlapping speech bubbles in brand colours "
     "with Arabic text inside. Question marks and lightbulb icons. Clean flat "
     "illustration on white.",
     "campaign/week-01-bold/posts/day-07-thread.md"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 2 — EDUCATE & DATA  |  Days 8–14  |  May 8–14, 2026
    # Topics: research & data, influencer mechanisms, historical advertising
    #         patterns, 46B MENA beauty market, body dysmorphia statistics,
    #         social media addiction research, Saudi Arabia body image study
    # ══════════════════════════════════════════════════════════════════════════

    # ── DAY 8 ─────────────────────────────────────────────────────────────────
    (8, "2026-05-08", "Week 2 – Educate", "Research & Data",
     "Instagram", "Carousel",
     "أحدث أبحاث تأثير السوشيال ميديا على المرأة",
     "مش رأي — ده بحث علمي. الأرقام بتتكلم.",
     "8-slide data carousel. Each slide = one key research finding: source name "
     "top-left in small MarkaziText, finding headline bold Rubik #ff009f, one-line "
     "explanation, sample size stat. #3db5aa (green) credibility accent bar on left "
     "edge. White card on light-grey background.",
     "Clean research-poster aesthetic: white cards on subtle grid paper background. "
     "Green accent bars, Arabic source citations, data bar charts in brand palette. "
     "Academic meets accessible.",
     "campaign/week-02-educate/posts/day-08-carousel.md"),

    (8, "2026-05-08", "Week 2 – Educate", "Research & Data",
     "Blog", "Blog Article",
     "الأبحاث العلمية وتأثير السوشيال ميديا على صورة المرأة",
     "ما تقوله الأبحاث عن السوشيال ميديا وصورة المرأة — والأرقام مفاجئة.",
     "2 200-word research-synthesis article. Covers 5 key peer-reviewed studies: "
     "body image deterioration, self-esteem correlation, purchase behaviour shifts, "
     "social comparison amplification, addiction-loop design. Full Arabic citations. "
     "Accessible language with expert quotes.",
     "Hero illustration: academic research papers and journals arranged artistically; "
     "social media platform icons overlaid as bookmarks. Green and pink colour "
     "palette. Bookish but modern editorial style.",
     "campaign/blog/article-03-research-social-media.md"),

    # ── DAY 9 ─────────────────────────────────────────────────────────────────
    (9, "2026-05-09", "Week 2 – Educate", "Research & Data",
     "Instagram", "Thread",
     "آليات الإنفلوانسر: كيف يعمل النظام من الداخل؟",
     "من العقد للمحتوى للضغط النفسي — رحلة المنتج من المصنع لعقلك.",
     "7-post educational thread. Post 1: brand-influencer contract structure. "
     "Post 2: disclosure laws vs. Arab market reality. Post 3: how authenticity is "
     "manufactured. Post 4: micro vs. macro influencer ROI for brands. Post 5: "
     "the 'relatability' illusion mechanics. Post 6: audience data sold to brands. "
     "Post 7: what you can do — awareness tools.",
     "Flowchart-series thread cards: each card has a process arrow or system diagram "
     "showing one step of the influencer machine. Money-flow arrows, brand boxes, "
     "audience funnel icons. Consistent card design, green + pink accent.",
     "campaign/week-02-educate/posts/day-09-thread.md"),

    (9, "2026-05-09", "Week 2 – Educate", "Research & Data",
     "Facebook", "Story Sequence",
     "خلف الكواليس: كيف تختار الشركات الإنفلوانسر؟",
     "العملية اللي ما بتشوفيهاش — وكيف تأثر عليكي.",
     "4-story behind-the-scenes sequence. Story 1: brand brief document mockup. "
     "Story 2: engagement-rate spreadsheet visual. Story 3: content approval process "
     "diagram. Story 4: 'هل تعرفتي على ده من قبل؟' poll sticker.",
     "Documentary-style mockups: contract papers, spreadsheets, brand brief "
     "documents layered with Instagram grid previews. Moody blue-pink lighting. "
     "Newspaper-print meets social-media aesthetic.",
     "campaign/week-02-educate/stories/day-09-story-sequence.md"),

    # ── DAY 10 ────────────────────────────────────────────────────────────────
    (10, "2026-05-10", "Week 2 – Educate", "Research & Data",
      "Instagram", "Infographic",
      "أنماط الإعلان التاريخية التي استهدفت المرأة: 1950–2026",
      "من الخمسينيات للألفية — الإعلان دايماً عارف ضعفك.",
      "Tall timeline infographic. 7 eras: 1950s (housewife perfection), 1970s "
      "(liberation paradox), 1980s (supermodel impossible standard), 1990s (heroin "
      "chic), 2000s (celebrity skin), 2010s (natural beauty fakeout), 2020s (body "
      "positivity weaponised). Each era: mini ad mockup + manipulation tactic label. "
      "Sepia to neon colour progression.",
      "Vertical vintage-to-modern timeline collage: top is sepia newspaper ad "
      "aesthetic, bottom is neon digital screen aesthetic. Each era section has a "
      "small representative ad mockup. Brand pink spine running the full height.",
      "campaign/week-02-educate/infographics/day-10-infographic.md"),

    (10, "2026-05-10", "Week 2 – Educate", "Research & Data",
      "Facebook", "Single Post",
      "الإعلان التاريخي وتأثيره على المرأة العربية",
      "إزاي السوق علّم المرأة العربية تكره جسمها — وإمتى بدأ ده؟",
      "Single Facebook educational image-post. Vintage Arabic magazine ad "
      "reimagined in modern graphic design. Long-form caption (200+ words) "
      "exploring advertising's historic role in shaping Arab women's body "
      "standards. Source citations included.",
      "Split composition: sepia-toned vintage Arabic beauty ad (left) mirrored by "
      "an equivalent contemporary social media ad (right). Contrast of era with "
      "identical manipulation message visible.",
      "campaign/week-02-educate/posts/day-10-single.md"),

    # ── DAY 11 ────────────────────────────────────────────────────────────────
    (11, "2026-05-11", "Week 2 – Educate", "Research & Data",
      "Instagram", "Carousel",
      "سوق الجمال في منطقة MENA: 46 مليار دولار",
      "46 مليار دولار — ومصر والسعودية في القلب.",
      "6-slide market-data carousel. Slide 1: MENA total market size hero stat. "
      "Slide 2: Egypt country breakdown with % share. Slide 3: Saudi Arabia "
      "breakdown. Slide 4: fastest-growing segments (skincare +18% YoY). Slide 5: "
      "key multinational players + Arab-owned brands emerging. Slide 6: what this "
      "means for Arab women consumers — power reframe.",
      "MENA region map with glowing data-point clusters over Egypt, Saudi, UAE. "
      "Bar charts in brand colours. Financial-report aesthetic blended with Arabic "
      "editorial typography. Night-map satellite style.",
      "campaign/week-02-educate/posts/day-11-carousel.md"),

    (11, "2026-05-11", "Week 2 – Educate", "Research & Data",
      "Blog", "Blog Article",
      "سوق الجمال في الشرق الأوسط: 46 مليار دولار وأنتِ في المنتصف",
      "من يستفيد فعلاً من سوق الجمال في الشرق الأوسط؟",
      "2 000-word market analysis. MENA beauty market structure, controlling "
      "multinationals, Arab-owned emerging brands, Egypt and Saudi Arabia as "
      "key markets, how Arab women are positioned as consumers vs. potential "
      "entrepreneurs, dokanelbanat.com as community counter-model.",
      "Hero: stylised MENA map constructed from beauty product icons (lipstick as "
      "building, foundation bottles as towers). Glowing green and pink overlay. "
      "Data-art editorial quality.",
      "campaign/blog/article-04-mena-beauty-market.md"),

    # ── DAY 12 ────────────────────────────────────────────────────────────────
    (12, "2026-05-12", "Week 2 – Educate", "Research & Data",
      "Instagram", "Story Sequence",
      "إحصائيات تشوه صورة الجسم (Body Dysmorphia) بين النساء",
      "الأرقام اللي محتاجة تعرفيها — ومش الصحافة بتتكلم فيها.",
      "5-story educational sequence. Story 1: global BDD prevalence stat on "
      "full-bleed soft background. Story 2: Arab-region specific data. Story 3: "
      "social media correlation coefficient (research). Story 4: age-of-onset "
      "statistics. Story 5: gentle CTA — dokanelbanat.com awareness resources "
      "link + 'إنتي مش لوحدك'.",
      "Clinical-meets-human illustration series: soft watercolour woman looking "
      "into a fractured mirror; statistics float as bold overlaid text. Green "
      "(#3db5aa) and pink (#ff009f) palette. Compassionate, not alarming.",
      "campaign/week-02-educate/stories/day-12-story-sequence.md"),

    (12, "2026-05-12", "Week 2 – Educate", "Research & Data",
      "Facebook", "Carousel",
      "تشوه صورة الجسم والشراء القهري: الصلة الخفية",
      "ما بين الوسواس والتسوق — الصلة اللي العلم أثبتها.",
      "5-slide carousel connecting BDD statistics to compulsive buying disorder. "
      "Cites peer-reviewed sources. Sensitive tone: factual, non-stigmatising. "
      "Each slide: research finding → real-life manifestation → awareness note. "
      "Professional mental-health messaging standards.",
      "Mosaic of compassionate data: gentle illustrated faces with embedded "
      "statistic numbers, pastel colour base with #ff009f accents. Informational "
      "and warm simultaneously.",
      "campaign/week-02-educate/posts/day-12-carousel.md"),

    # ── DAY 13 ────────────────────────────────────────────────────────────────
    (13, "2026-05-13", "Week 2 – Educate", "Research & Data",
      "Instagram", "Infographic",
      "بحث إدمان السوشيال ميديا: الخوارزمية صُممت ضدك",
      "الخوارزمية مش بتوصيلك محتوى — بتوصيلك إحساس.",
      "Tall portrait infographic. Five sections with icons: (1) Dopamine Loop "
      "diagram, (2) Variable Reward Schedule (slot-machine comparison), (3) "
      "Infinite Scroll design intent, (4) Comparison-content amplification data, "
      "(5) Attention economy dollar value per user. Each section cites a research "
      "source. Dark background, neon accents.",
      "Brain anatomy illustration with social-media platform logos as neural-pathway "
      "nodes. Dopamine molecules rendered as small pink bubbles travelling between "
      "nodes. Deep dark background, neon pink and green accents. Scientific editorial.",
      "campaign/week-02-educate/infographics/day-13-infographic.md"),

    (13, "2026-05-13", "Week 2 – Educate", "Research & Data",
      "Facebook", "Thread",
      "إدمان السوشيال ميديا: هل أنتِ مدمنة؟ 7 أسئلة للتقييم",
      "7 أسئلة تحددي بيهم علاقتك الحقيقية مع السوشيال ميديا.",
      "Facebook thread: root post introduces the 7-question self-assessment. "
      "Each question posted as a separate comment with brief 2-sentence explanation "
      "of why it matters. Final comment: free PDF from dokanelbanat.com with full "
      "assessment + scoring guide.",
      "Assessment worksheet aesthetic: question card design, bold Arabic numbers, "
      "#ff009f circle bullets, clean white background. Clinical but warm. Could be "
      "screenshot-shared as standalone card.",
      "campaign/week-02-educate/posts/day-13-thread.md"),

    # ── DAY 14 ────────────────────────────────────────────────────────────────
    (14, "2026-05-14", "Week 2 – Educate", "Research & Data",
      "Instagram", "Carousel",
      "دراسة السعودية: صورة الجسم والشبكات الاجتماعية عند المرأة العربية",
      "دراسة سعودية تكشف: الفتيات العربيات أكتر تأثراً — وده السبب.",
      "7-slide carousel presenting Saudi Arabia body-image study. Slide 1: study "
      "headline and key finding. Slides 2–5: Arab-specific cultural factors — "
      "modesty-culture paradox, family comparison dynamics, wedding culture body "
      "pressure, skin-lightening product demand. Slide 6: Egypt comparison data. "
      "Slide 7: Week 2 recap + Week 3 teaser.",
      "Research-presentation style: slide cards with Saudi cultural imagery (modest "
      "dress, family gathering setting) intersected with data charts. Warm tan and "
      "#ff009f colour palette. Academic but culturally grounded.",
      "campaign/week-02-educate/posts/day-14-carousel.md"),

    (14, "2026-05-14", "Week 2 – Educate", "Research & Data",
      "Blog", "Blog Article",
      "الدراسة السعودية: كيف تؤثر وسائل التواصل على صورة الجسم في العالم العربي",
      "ما كشفته الدراسات العربية عن صورة الجسم — والأرقام مقلقة.",
      "2 100-word article synthesising key Arab-region body-image studies. Focus "
      "on Saudi Arabia research, Egypt comparison data, cultural specificities of "
      "beauty pressure (modesty paradox, wedding culture, family dynamics), "
      "whitening-product market as indicator. Closes with dokanelbanat.com awareness "
      "framework introduction ahead of Week 3.",
      "Hero: illustrated map of Arab world with embedded data-point annotations. "
      "Female silhouettes at key geographic markers. Research-journal aesthetic "
      "with warm green and pink gradient overlay.",
      "campaign/blog/article-05-saudi-body-image-study.md"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 3 — SHIFT  |  Days 15–21  |  May 15–21, 2026
    # Topics: mindset shift, minimalism, daily aware routine, self-sufficiency,
    #         4 pillars, clean beauty philosophy, real beauty standards
    # ══════════════════════════════════════════════════════════════════════════

    # ── DAY 15 ────────────────────────────────────────────────────────────────
    (15, "2026-05-15", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "بداية التحول: اخترنا الوعي",
      "الوعي مش قرار لحظة — هو عادة يومية. وبتبدأ هنا.",
      "6-slide pivot carousel. Visual tone shifts: lighter backgrounds (#fff5fb), "
      "airy typography. Slide 1: 'أسبوع التحول' title card — sunrise colour gradient. "
      "Slide 2: what the shift means (not restriction, but awareness). Slides 3–5: "
      "three mindset reframes from Weeks 1–2. Slide 6: introduces 4 Pillars "
      "framework — preview graphic.",
      "Sunrise aesthetic: Arab woman stretching in morning outdoor light. Brand "
      "colour sunrise gradient (#ffbe3d → #ff6b4a → #ff009f). Clean, airy, hopeful. "
      "Rubik Bold Arabic text as motivational overlay.",
      "campaign/week-03-shift/posts/day-15-carousel.md"),

    (15, "2026-05-15", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Single Post",
      "التحول نحو الاستهلاك الواعي: من أين تبدئين؟",
      "إزاي تبدأي رحلة الوعي من غير ما تحسي بذنب على الماضي.",
      "Single warm-tone Facebook post. Long-form caption (180 words): non-judgmental "
      "reflection on past unconscious spending, reframe as learning, 3 concrete "
      "first steps to begin the conscious journey. Personal, empathetic brand voice.",
      "Warm flat-lay: simple journal open to blank page, 3 minimal skincare products, "
      "morning coffee, soft natural side-lighting. Cream and blush tones. "
      "Aspirational-real.",
      "campaign/week-03-shift/posts/day-15-single.md"),

    # ── DAY 16 ────────────────────────────────────────────────────────────────
    (16, "2026-05-16", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Story Sequence",
      "فلسفة المينيماليزم في الجمال: أقل يساوي أكثر",
      "أقل = أكتر. فلسفة المينيماليزم في الجمال والحياة اليومية.",
      "5-story vertical sequence. Story 1: define beauty minimalism vs. deprivation. "
      "Story 2: capsule skincare concept (5-product routine). Story 3: capsule "
      "wardrobe concept (10 pieces). Story 4: 5-question routine audit ('كام منتج "
      "عندك دلوقتي؟'). Story 5: CTA to dokanelbanat.com minimalism guide.",
      "Minimal aesthetic stories: white and cream full-bleed. Simple product "
      "groupings flat-lay. 'less is more' typography treatment. Clean geometric "
      "framing. Each story = one concise message.",
      "campaign/week-03-shift/stories/day-16-story-sequence.md"),

    (16, "2026-05-16", "Week 3 – Shift", "Alternative Mindset",
      "Blog", "Blog Article",
      "المينيماليزم والجمال: كيف تبنين روتيناً واعياً بخمسة منتجات فقط",
      "الروتين الواعي مش محتاج 20 منتج — 5 كفاية.",
      "1 900-word minimalist beauty guide. Sections: (1) capsule skincare philosophy, "
      "(2) ingredient awareness basics, (3) multi-use products guide, (4) slow beauty "
      "movement explained, (5) Arab skin types mini-guide, (6) dokanelbanat.com curated "
      "5-product starter list. Practical and educational.",
      "Hero: flat-lay of exactly 5 skincare products on white marble, soft directional "
      "natural light, clean cream and pink tones. Minimalist editorial photography. "
      "Intentional negative space.",
      "campaign/blog/article-06-minimalism-beauty.md"),

    # ── DAY 17 ────────────────────────────────────────────────────────────────
    (17, "2026-05-17", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "الروتين اليومي الواعي: خطوة بخطوة",
      "يومك من غير ما السوشيال ميديا تحكمه — روتين يومي واعي.",
      "7-slide practical carousel. Each slide = time block: Morning Ritual, "
      "Mid-Morning Check-in, Lunch Mindfulness, Afternoon Recharge, Evening Wind-down, "
      "Night Reflection, Weekly Review. Each block: awareness check question, "
      "intentional choice suggestion, gratitude micro-practice.",
      "Day-planner illustration: time blocks as clean calendar rows with warm "
      "icons. #ffbe3d for morning, #3db5aa for afternoon, #ff009f for evening. "
      "Planner-book page texture background. Organised, calming aesthetic.",
      "campaign/week-03-shift/posts/day-17-carousel.md"),

    (17, "2026-05-17", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Story Sequence",
      "أول 30 دقيقة في يومك: لماذا هي الأهم؟",
      "أول 30 دقيقة في يومك — اللي بتقرري فيها مزاجك كله.",
      "4-story practical guide. Story 1: research-backed reason why first 30 minutes "
      "shape the entire day. Stories 2–3: phone-free morning steps (with alternative "
      "habit suggestions). Story 4: 7-day morning challenge invitation with "
      "dokanelbanat.com community link.",
      "Morning photography style stories: sunrise light through curtain, open "
      "journal, glass of water, no phone present. Warm cream and gold tones. "
      "Calm, inviting, peaceful atmosphere.",
      "campaign/week-03-shift/stories/day-17-story-sequence.md"),

    # ── DAY 18 ────────────────────────────────────────────────────────────────
    (18, "2026-05-18", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Thread",
      "الاكتفاء الذاتي: المرأة القوية التي تكتفي بنفسها",
      "الاكتفاء مش عزلة — هو قوة. الفرق بين الاثنين.",
      "8-post thread. Post 1: define self-sufficiency vs. isolation. Post 2: "
      "financial self-sufficiency basics. Post 3: emotional self-sufficiency (not "
      "seeking external validation). Post 4: knowledge self-sufficiency (own your "
      "learning). Post 5: skill-building as asset. Post 6: community vs. dependency "
      "distinction. Post 7: Arab cultural context (family vs. self). Post 8: "
      "dokanelbanat.com as resource hub for self-sufficient women.",
      "Thread card series: each card — bold Arabic keyword (الاكتفاء / القوة / "
      "المعرفة etc.) centred, supporting stat or insight below, #ff009f and "
      "#3db5aa alternating accent lines. Strong, clean visual identity.",
      "campaign/week-03-shift/posts/day-18-thread.md"),

    (18, "2026-05-18", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Infographic",
      "أربعة أركان المرأة الواعية: النموذج الكامل",
      "4 أركان — 1 هوية. اكتشفي النموذج اللي بيبنيكي.",
      "Square 4-quadrant infographic (1080×1080). Top-left: Self-Sufficiency "
      "الاكتفاء الذاتي (#ff009f). Top-right: Conscious Awareness الوعي الاستهلاكي "
      "(#ff6b4a). Bottom-left: Intentional Routine الروتين المتعمد (#ffbe3d). "
      "Bottom-right: Empowering Community المجتمع الداعم (#3db5aa). Each quadrant: "
      "pillar icon + Arabic name + 3-word description.",
      "Four-quadrant geometric design: each section a distinct brand colour with "
      "centred icon and Arabic label. Clean sans-serif type. Symmetrical, balanced, "
      "memorable visual framework. White divider lines between quadrants.",
      "campaign/week-03-shift/infographics/day-18-infographic.md"),

    # ── DAY 19 ────────────────────────────────────────────────────────────────
    (19, "2026-05-19", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "فلسفة الجمال النظيف: Clean Beauty بالعربي",
      "Clean Beauty مش ترند — هو حق. المرأة العربية والمكونات النظيفة.",
      "6-slide carousel. Slide 1: clean beauty definition in Arab cultural context. "
      "Slides 2–3: 6 common harmful ingredients (Arabic INCI names) with red-flag "
      "icons and safe alternatives. Slide 4: Arab-specific skin concerns and "
      "clean solutions. Slide 5: how to read an Arabic product label. Slide 6: "
      "dokanelbanat.com clean-pick recommendations.",
      "Clean beauty product flat-lay series: white marble, fresh botanical elements "
      "(aloe, rose petals, argan), Arabic ingredient label visible. #3db5aa green "
      "for 'clean' items, red for 'avoid'. Scientific-beautiful aesthetic.",
      "campaign/week-03-shift/posts/day-19-carousel.md"),

    (19, "2026-05-19", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Thread",
      "كيف تقرئين مكونات أي منتج تجميل بالعربي؟",
      "قبل ما تشتري أي منتج تجميل — اقرأي ده الأول.",
      "Facebook thread: step-by-step guide to reading product ingredient lists. "
      "Root post explains why it matters. 6 comments: (1) INCI name basics, "
      "(2) order of concentration, (3) 3 always-avoid ingredients, (4) 3 "
      "look-for ingredients, (5) Arabic vs. INCI name reference table, (6) "
      "dokanelbanat.com ingredient checker tool link.",
      "Thread visual cards: product ingredient list mockup with highlighted "
      "warning ingredients in red, green-tick safe alternatives. Magnifying glass "
      "illustration. Clean educational flat design.",
      "campaign/week-03-shift/posts/day-19-thread.md"),

    # ── DAY 20 ────────────────────────────────────────────────────────────────
    (20, "2026-05-20", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Story Sequence",
      "معايير الجمال الحقيقية: من قال إنها صح؟",
      "من قرر إن الشعر الناعم أحسن؟ ومن قرر إن البشرة الفاتحة أجمل؟",
      "5-story deconstruction sequence. Story 1: 'من قرر؟' — powerful opening "
      "question. Story 2: colonial history of Arab beauty standards (brief). "
      "Story 3: diversity showcase — 4 types of natural Arab beauty (skin tone, "
      "hair texture, body type, features). Story 4: media's standard-manufacturing "
      "role. Story 5: empowerment reframe + 'إيه رأيك؟' question sticker.",
      "Diverse Arab women portrait series (illustrated): each portrait different "
      "natural feature set. Each has a deconstructed beauty standard as crossed-out "
      "overlay text. Vibrant colours per portrait. Empowering, celebratory.",
      "campaign/week-03-shift/stories/day-20-story-sequence.md"),

    (20, "2026-05-20", "Week 3 – Shift", "Alternative Mindset",
      "Blog", "Blog Article",
      "معايير الجمال العربية: من صنعها ولماذا؟",
      "الجمال العربي تحت المجهر — من وضع المعايير ولمصلحة من؟",
      "2 000-word article examining Arab beauty standards. Covers: colonial-era "
      "imposition of Eurocentric standards, media's ongoing manufacturing role, "
      "skin-lightening industry ($8.6B globally) as indicator, natural Arab beauty "
      "diversity (geography-based), celebration framework for reclaiming standards. "
      "Rich with historical and commercial research.",
      "Hero: illustrated collage of diverse Arab women in natural everyday settings "
      "(not fashion). Warm earth tones, deconstructed beauty-product imagery as "
      "background texture. Empowering editorial quality.",
      "campaign/blog/article-07-arab-beauty-standards.md"),

    # ── DAY 21 ────────────────────────────────────────────────────────────────
    (21, "2026-05-21", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "ملخص الأسبوع الثالث: التحول اكتمل",
      "3 أسابيع — والتحول بدأ. إيه اللي تغير في طريقة تفكيرك؟",
      "6-slide recap. Lighter colour tone reflecting the shift. Slide 1: 'Week 3 "
      "Shift Complete' title card with airy gradient. Slides 2–5: one insight per "
      "shift topic (minimalism, 4 pillars, clean beauty, real standards). Slide 6: "
      "Week 4 Empower teaser with bold 'أنتِ مكتملة' preview. Community engagement "
      "CTA.",
      "Mosaic of Week 3 content thumbnails in lighter colour treatment. Airy, open "
      "aesthetic. Brand colours in softer tones. Bold Week 4 teaser slide as final "
      "card with energising typography.",
      "campaign/week-03-shift/posts/day-21-carousel.md"),

    (21, "2026-05-21", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Infographic",
      "قبل الوعي وبعده: جدول المقارنة الكامل",
      "حياتك قبل الوعي الاستهلاكي وبعده — الفرق واضح.",
      "Before/After comparison infographic (Facebook landscape format). Two columns: "
      "قبل الوعي (dark, cluttered visual) vs بعد الوعي (clean, spacious visual). "
      "8 comparison rows: spending habits, social media time, self-image, buying "
      "triggers, community quality, knowledge level, daily routine, brand loyalty. "
      "Each row with icon.",
      "Split-table design: left half dark/chaotic visual language, right half "
      "clean/calm. Same Arab woman illustrated in both halves but in contrasting "
      "energy states. Brand colour transformation story visible.",
      "campaign/week-03-shift/infographics/day-21-infographic.md"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 4 — EMPOWER  |  Days 22–30  |  May 22–30, 2026
    # Topics: empowerment, community, women entrepreneurs, dokanelbanat.com
    #         ecosystem, loyal community building, conscious consuming,
    #         brand building for women
    # ══════════════════════════════════════════════════════════════════════════

    # ── DAY 22 ────────────────────────────────────────────────────────────────
    (22, "2026-05-22", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "أنتِ مكتملة: البيان الأول",
      "مش محتاجة أي منتج علشان تبقي مكتملة. أنتِ مكتملة دلوقتي.",
      "7-slide manifesto carousel. Maximum typographic impact — no product imagery. "
      "Slide 1: 'أنتِ مكتملة' full-bleed bold statement on #ff009f. Slides 2–6: "
      "each completeness dimension (inner strength, knowledge, authentic beauty, "
      "community, purpose). Slide 7: dokanelbanat.com as home for complete women. "
      "Rubik Extra Bold throughout.",
      "Power-poster typographic design: full-frame bold Arabic text, each slide "
      "a different brand gradient (pink → orange → yellow → green → back to pink). "
      "No photographs — pure type and colour. Manifesto energy.",
      "campaign/week-04-empower/posts/day-22-carousel.md"),

    (22, "2026-05-22", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Single Post",
      "رسالة dokanelbanat.com: لماذا أنشأنا هذه المنصة",
      "السبب الحقيقي وراء dokanelbanat.com — مش متجر. مجتمع.",
      "Founder's-voice Facebook post. Long-form narrative (250 words) telling the "
      "origin story of dokanelbanat.com: the problem observed, the gap in the market, "
      "the community vision, what makes it different. Personal, authentic, mission-led. "
      "Invites community to share their 'why'.",
      "Warm editorial portrait: woman at desk with open journal and laptop, "
      "dokanelbanat.com logo subtle on screen, candid natural light, #fff5fb tones. "
      "Real, not staged.",
      "campaign/week-04-empower/posts/day-22-single.md"),

    # ── DAY 23 ────────────────────────────────────────────────────────────────
    (23, "2026-05-23", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Story Sequence",
      "قصص من مجتمعنا: نساء اخترن الوعي",
      "مش بنحكي عنهم — بنحكي معاهم. قصص حقيقية من مجتمعنا.",
      "5-story community spotlight. Each story = one community member (anonymised "
      "or with consent): her challenge described in 1 line, the shift moment, "
      "the outcome. Story 5: call for submissions — 'شاركي قصتك مع dokanelbanat.com'.",
      "Portrait spotlight: soft-vignette illustrated portrait per story, first-person "
      "Arabic quote in pull-quote style, #ff009f accent frame. Community member "
      "name (or alias) in Rubik Bold. Warm, human.",
      "campaign/week-04-empower/stories/day-23-story-sequence.md"),

    (23, "2026-05-23", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "نساء رائدات: قصص من مجتمع dokanelbanat.com",
      "نساء عربيات اخترن الوعي — قصصهم بتغير مجتمعات.",
      "1 800-word community spotlight article. 4–5 stories of Arab women "
      "(entrepreneurs and conscious consumers) from the dokanelbanat.com community. "
      "Each story: background, challenge, mindset shift catalyst, measurable change. "
      "dokanelbanat.com positioned as community catalyst, not just platform.",
      "Hero: illustrated group portrait of 4–5 diverse Arab women in professional "
      "natural settings. Warm community energy, not corporate. Editorial quality "
      "with brand colour accents.",
      "campaign/blog/article-08-community-women-stories.md"),

    # ── DAY 24 ────────────────────────────────────────────────────────────────
    (24, "2026-05-24", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Infographic",
      "النساء رائدات الأعمال في العالم العربي: أرقام الإلهام",
      "المرأة العربية رائدة أعمال — والأرقام بتقول حاجة كبيرة.",
      "Portrait infographic on Arab women entrepreneurs. Stats: % of MENA startups "
      "founded by women, Egypt female entrepreneurship growth rate, top sectors "
      "(e-commerce, beauty, education, food), key barriers (access to capital, "
      "networking gaps), dokanelbanat.com as support ecosystem. Bold data, "
      "empowering framing.",
      "Bold entrepreneurship infographic: rising arrow charts, female silhouette "
      "with business elements (laptop, storefront, product). #ff009f and #ffbe3d "
      "dominant. Arabic stat callouts. Confident, aspirational visual tone.",
      "campaign/week-04-empower/infographics/day-24-infographic.md"),

    (24, "2026-05-24", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Carousel",
      "كيف تبدئين مشروعك الصغير من الصفر؟",
      "من الفكرة للبيعة الأولى — دليل عملي للمرأة العربية.",
      "6-slide practical startup carousel. Steps: (1) idea validation (3 questions), "
      "(2) minimal viable product concept, (3) first sales channel options, "
      "(4) pricing psychology basics, (5) building your first audience, (6) how "
      "dokanelbanat.com fits into every step. Actionable and specific.",
      "Step-by-step guide cards: numbered steps 1–6, icon illustration per step, "
      "#ff6b4a warm orange dominant, professional but approachable. Small business "
      "imagery (handmade product, phone camera, simple storefront).",
      "campaign/week-04-empower/posts/day-24-carousel.md"),

    # ── DAY 25 ────────────────────────────────────────────────────────────────
    (25, "2026-05-25", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Thread",
      "منظومة dokanelbanat.com: كل ما تحتاجينه في مكان واحد",
      "واحدة سألتني: إيه اللي عنده dokanelbanat.com فعلاً؟ — اتفضلي.",
      "8-post thread. Each post = one dokanelbanat.com ecosystem pillar: (1) "
      "awareness content library, (2) educational resources, (3) community forum, "
      "(4) curated conscious product discovery, (5) entrepreneur support programme, "
      "(6) expert network access, (7) events and workshops, (8) weekly newsletter. "
      "Each post: pillar name → what you get → how to access.",
      "Ecosystem map thread cards: hub-and-spoke diagram — dokanelbanat.com logo "
      "at centre, each pillar as a labelled spoke. Consistent card design. Brand "
      "colour per pillar. Final card is full ecosystem map.",
      "campaign/week-04-empower/posts/day-25-thread.md"),

    (25, "2026-05-25", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "منظومة dokanelbanat.com: كيف تستفيدين من كل خدمة؟",
      "كل ما تقدمه dokanelbanat.com — وكيف تستخدمينه لصالحك.",
      "2 000-word complete ecosystem guide. Each section = one platform offering "
      "with specific use-cases, how to access, real benefit for Arab women. "
      "Written as a practical onboarding guide for new members. SEO-optimised "
      "for dokanelbanat.com brand keywords.",
      "Hero: dokanelbanat.com interface mockup surrounded by community icons, "
      "educational elements, curated product imagery. Brand-accurate colour palette. "
      "Editorial software-screenshot aesthetic.",
      "campaign/blog/article-09-dokanelbanat-ecosystem.md"),

    # ── DAY 26 ────────────────────────────────────────────────────────────────
    (26, "2026-05-26", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "بناء مجتمع مخلص: الفرق بين الجمهور والمجتمع",
      "مليون فولوور مش مجتمع — وألف امرأة واعية أقوى منهم.",
      "7-slide community-building carousel. Slide 1: audience vs. community "
      "distinction (powerful visual contrast). Slides 2–5: 4 pillars of a real "
      "community (shared values, two-way dialogue, mutual support, collective "
      "growth). Slide 6: dokanelbanat.com community principles. Slide 7: join CTA "
      "with what to expect in first 7 days as member.",
      "Community visualisation: concentric-circle diagram showing depth (community) "
      "vs breadth (audience). Warm illustrated human figures connected by lines. "
      "#3db5aa green dominant for community slides. Conversations and connections "
      "depicted.",
      "campaign/week-04-empower/posts/day-26-carousel.md"),

    (26, "2026-05-26", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Story Sequence",
      "كيف تنضمين لمجتمع dokanelbanat.com؟",
      "3 خطوات بس — وتبقي جزء من أقوى مجتمع نساء عربيات.",
      "3-story onboarding sequence. Story 1: 'ما الذي ستحصلين عليه' — member "
      "benefits visual. Story 2: step-by-step joining guide with screen mockup. "
      "Story 3: your first community action — 'عرّفي عن نفسك' with dokanelbanat.com "
      "community link.",
      "Onboarding welcome design: open-door visual metaphor (warm glow inside). "
      "Step icons clean and minimal. Community member avatar cluster illustration. "
      "Welcoming brand colour gradient.",
      "campaign/week-04-empower/stories/day-26-story-sequence.md"),

    # ── DAY 27 ────────────────────────────────────────────────────────────────
    (27, "2026-05-27", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Story Sequence",
      "الاستهلاك الواعي في 30 يوم: رحلتنا معاً",
      "30 يوم من الوعي — والرحلة مش خلصت، هي بس بدأت.",
      "5-story campaign retrospective. Story 1: campaign arc overview (4 weeks "
      "visual map). Story 2: most impactful Week 1 reveal + community reaction. "
      "Story 3: most shared Week 2 data stat. Story 4: most resonant Week 3 "
      "mindset shift moment. Story 5: Week 4 empowerment highlight + 'ما الذي "
      "تغير لديك؟' question sticker.",
      "Film-strip retrospective: each story shows mini-thumbnails of campaign "
      "highlights in that phase. Nostalgic-meets-forward-looking aesthetic. "
      "Brand gradient evolution: dark (Week 1) → light (Week 4).",
      "campaign/week-04-empower/stories/day-27-story-sequence.md"),

    (27, "2026-05-27", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Thread",
      "30 يوم من الوعي: 10 دروس كبيرة تعلمناها معاً",
      "كل يوم علمنا حاجة — دي أهم 10 دروس من 30 يوم.",
      "Facebook thread: 10 lessons from the campaign. Root post frames the "
      "retrospective. Each of 10 comments = one lesson: lesson statement, "
      "the Week/Day it came from, community engagement data point that validated "
      "it, forward implication for daily life.",
      "Campaign learning cards: numbered lessons in clean list card format. "
      "#ff009f bullet number circles. Shareable as screenshot series. Summary "
      "design at high-contrast for readability.",
      "campaign/week-04-empower/posts/day-27-thread.md"),

    # ── DAY 28 ────────────────────────────────────────────────────────────────
    (28, "2026-05-28", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "بناء البراند الشخصي للمرأة العربية",
      "براندك الشخصي مش لوجو ولا كولر — هو قيمتك في السوق.",
      "6-slide personal branding carousel. Slide 1: redefine personal brand for "
      "Arab women context. Slide 2: values-based positioning (not just aesthetics). "
      "Slide 3: authentic storytelling framework. Slide 4: LinkedIn for Arab "
      "professional women — quick guide. Slide 5: consistency over virality. "
      "Slide 6: dokanelbanat.com community as brand amplifier.",
      "Personal brand builder carousel: clean professional design, woman-at-work "
      "illustrated, brand identity elements depicted (colour swatches, font samples, "
      "content calendar). #ff009f accent. Empowering business aesthetic.",
      "campaign/week-04-empower/posts/day-28-carousel.md"),

    (28, "2026-05-28", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "البراند الشخصي للمرأة العربية: دليل البداية في 2026",
      "كيف تبنين براند شخصي أصيل كامرأة عربية في 2026؟",
      "2 100-word personal branding guide. Sections: (1) discovering unique "
      "positioning, (2) values audit exercise, (3) building authentic online "
      "presence on LinkedIn and Instagram, (4) Arab cultural context in "
      "professional branding, (5) community leverage strategy, (6) 90-day "
      "brand-building action plan. Practical and specific to Arab women context.",
      "Hero: confident Arab professional woman at whiteboard with brand strategy "
      "elements sketched behind her. Warm office setting. Brand colours integrated "
      "into environment design. Inspirational, real.",
      "campaign/blog/article-10-personal-brand-arab-women.md"),

    # ── DAY 29 ────────────────────────────────────────────────────────────────
    (29, "2026-05-29", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Thread",
      "دليل الاستهلاك الواعي الذي يبقى معكِ للأبد",
      "مش حملة — طريقة حياة. الدليل الكامل للاستهلاك الواعي.",
      "8-post evergreen thread. Post 1: monthly spending audit protocol. Post 2: "
      "social media feed curation guide. Post 3: product ingredient awareness "
      "checklist. Post 4: community validation before purchase habit. Post 5: "
      "supporting women-owned brands framework. Post 6: daily gratitude-body "
      "practice. Post 7: quarterly mindset review process. Post 8: annual "
      "brand-and-media detox guide.",
      "Evergreen handbook thread cards: each card styled as a 'chapter' in a "
      "clean handbook. Chapter number bold, #3db5aa green for 'permanent' feel. "
      "Practical icons, Arabic chapter titles. Could work as printable cards.",
      "campaign/week-04-empower/posts/day-29-thread.md"),

    (29, "2026-05-29", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Infographic",
      "خريطة قرار الشراء الواعي: 7 خطوات قبل أي منتج",
      "كل قرار شراء بيمر بـ7 خطوات — هل بتمشي عليهم؟",
      "Tall decision-flowchart infographic. 7-step purchase decision tree: "
      "(1) هل أحتاجه فعلاً؟ (2) بحث المكونات (3) مراجعة المجتمع (4) توافق "
      "القيم (5) الميزانية (6) التأثير البيئي (7) اشتري أو انتظري. "
      "Yes/No branches. Shareable poster format for Facebook.",
      "Flowchart poster: clean Arabic decision tree. Yes branches in #3db5aa "
      "green, pause branches in #ff009f, research branches in #ffbe3d. White "
      "background, clear typography, infographic-quality design.",
      "campaign/week-04-empower/infographics/day-29-infographic.md"),

    # ── DAY 30 ────────────────────────────────────────────────────────────────
    (30, "2026-05-30", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "يوم 30: أنتِ مكتملة — البيان النهائي",
      "30 يوم، ومجتمع كامل. أنتِ مكتملة وجاهزة تغيري الوعي.",
      "8-slide campaign finale. Slide 1: bold finale statement on #ff009f. "
      "Slides 2–4: 3-phase journey recap (Expose → Educate → Shift). Slide 5: "
      "community achievement metrics. Slide 6: 4 Pillars visual (final version). "
      "Slide 7: what comes next from dokanelbanat.com. Slide 8: closing manifesto "
      "'أنتِ مكتملة. دلوقتي وللأبد.' — most powerful design of the campaign.",
      "Grand finale carousel: campaign's strongest visual execution. Full-bleed "
      "brand gradients, #ffbe3d gold celebration accents, community portrait "
      "montage slide, closing manifesto in maximum typographic weight. Cinematic "
      "campaign-conclusion energy.",
      "campaign/week-04-empower/posts/day-30-carousel.md"),

    (30, "2026-05-30", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "30 يوم من الوعي: الرسالة الختامية لمجتمع dokanelbanat.com",
      "الرسالة الأخيرة — وبداية كل شيء.",
      "2 500-word campaign closing manifesto article. Full journey narrative: "
      "(1) what we exposed in Week 1, (2) what the data confirmed in Week 2, "
      "(3) the mindset shift we built in Week 3, (4) the community we became in "
      "Week 4. Closes with dokanelbanat.com long-term commitments to the community. "
      "Manifesto-quality writing — designed to be shared and bookmarked.",
      "Hero: wide community illustration — circle of diverse Arab women, sunrise "
      "behind them, empowering atmosphere. All brand colours harmonised. Cinematic "
      "campaign-finale quality. Designed to inspire.",
      "campaign/blog/article-11-campaign-closing-manifesto.md"),

    (30, "2026-05-30", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Story Sequence",
      "شكراً — وما الذي يأتي بعد ذلك مع dokanelbanat.com",
      "الحملة خلصت — المجتمع لأ. إيه اللي بيجي بعد كده؟",
      "4-story closing sequence. Story 1: campaign thank-you with key stats "
      "(reach / engagement / community growth). Story 2: what comes next from "
      "dokanelbanat.com (upcoming content, events, tools). Story 3: how to stay "
      "connected (newsletter subscribe, community join, social follow). Story 4: "
      "'شاركي تحولك' — share your transformation CTA with dokanelbanat.com hashtag.",
      "Closing celebration stories: confetti-particle effect in brand colours. "
      "Warm gratitude aesthetic. Forward-looking imagery (horizon, open door). "
      "Newsletter CTA card as final frame — clean and compelling.",
      "campaign/week-04-empower/stories/day-30-story-sequence.md"),
]

# ── Workbook Setup ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "30-Day Calendar"

# ── Styles ────────────────────────────────────────────────────────────────────
hdr_fill     = PatternFill("solid", fgColor=C_PRIMARY)
hdr_font     = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
alt_fill     = PatternFill("solid", fgColor=C_ALT)
plain_fill   = PatternFill("solid", fgColor=C_WHITE)
center_al    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_al      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin         = Side(style="thin", color="DDDDDD")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

# Week-band accent fills (very light)
wfill = {
    "Week 1 – Bold":    PatternFill("solid", fgColor="FFE5F5"),
    "Week 2 – Educate": PatternFill("solid", fgColor="E5F7FF"),
    "Week 3 – Shift":   PatternFill("solid", fgColor="E5FFE8"),
    "Week 4 – Empower": PatternFill("solid", fgColor="FFF8E5"),
}

# Platform colour fonts for col 5
platform_colours = {
    "Instagram": "C13584",
    "Facebook":  "1877F2",
    "Blog":      "FF6B4A",
}

# Content-type colour fonts for col 6
ctype_colours = {
    "Carousel":       "FF009F",
    "Single Post":    "333333",
    "Story Sequence": "FF2D55",
    "Infographic":    "3DB5AA",
    "Thread":         "FF6B4A",
    "Blog Article":   "FFBE3D",
}

# ── Header Row ────────────────────────────────────────────────────────────────
ws.append(HEADERS)
ws.row_dimensions[1].height = 32
for ci in range(1, len(HEADERS) + 1):
    c = ws.cell(row=1, column=ci)
    c.fill      = hdr_fill
    c.font      = hdr_font
    c.alignment = center_al
    c.border    = border

# ── Data Rows ─────────────────────────────────────────────────────────────────
for ri, row_data in enumerate(ROWS, start=2):
    full_row = list(row_data) + ["Ready to Publish"]
    ws.append(full_row)
    ws.row_dimensions[ri].height = 60

    week_key  = row_data[2]
    is_alt    = (ri % 2 == 0)
    base_fill = wfill.get(week_key, plain_fill)
    row_fill  = alt_fill if is_alt else base_fill

    platform  = row_data[4]
    ctype     = row_data[5]

    for ci in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.fill   = row_fill
        cell.border = border

        # Alignment
        if ci in (1, 2, 3, 5, 6, 12):
            cell.alignment = center_al
        else:
            cell.alignment = left_al

        # Special fonts
        if ci == 5:   # Platform
            col = platform_colours.get(platform, "222222")
            cell.font = Font(name="Calibri", color=col, bold=True, size=10)
        elif ci == 6: # Content Type
            col = ctype_colours.get(ctype, "222222")
            cell.font = Font(name="Calibri", color=col, bold=True, size=10)
        elif ci == 12: # Status
            cell.font = Font(name="Calibri", color="3DB5AA", bold=True, size=10)
        elif ci == 1:  # Day number
            cell.font = Font(name="Calibri", bold=True, size=12, color=C_PRIMARY)
        else:
            cell.font = Font(name="Calibri", size=10)

# ── Freeze Pane ───────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Column Widths ─────────────────────────────────────────────────────────────
WIDTHS = {
    1: 6,    # Day
    2: 14,   # Date
    3: 17,   # Week
    4: 22,   # Theme
    5: 13,   # Platform
    6: 17,   # Content Type
    7: 38,   # Topic (Arabic)
    8: 45,   # Hook Line
    9: 62,   # Design Brief Summary
    10: 58,  # Image Prompt Summary
    11: 52,  # File Path
    12: 18,  # Status
}
for ci, col_cells in enumerate(ws.columns, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(ci, 20)

# ── Summary Sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Campaign Summary")

type_counts = {}
week_counts = {}
platform_counts = {}
for r in ROWS:
    type_counts[r[5]]     = type_counts.get(r[5], 0) + 1
    week_counts[r[2]]     = week_counts.get(r[2], 0) + 1
    platform_counts[r[4]] = platform_counts.get(r[4], 0) + 1

summary = [
    ["dokanelbanat.com — 30-Day Campaign Calendar", ""],
    ["", ""],
    ["Metric", "Value"],
    ["Total Content Rows", len(ROWS)],
    ["Campaign Days", 30],
    ["Start Date", "2026-05-01"],
    ["End Date", "2026-05-30"],
    ["Platforms", ", ".join(sorted(platform_counts.keys()))],
    ["Status", "All Ready to Publish"],
    ["", ""],
    ["CONTENT TYPE", "COUNT"],
]
for k, v in sorted(type_counts.items()):
    summary.append([k, v])
summary += [["", ""], ["WEEK", "ROWS"]]
for k, v in sorted(week_counts.items()):
    summary.append([k, v])
summary += [["", ""], ["PLATFORM", "ROWS"]]
for k, v in sorted(platform_counts.items()):
    summary.append([k, v])

for sr in summary:
    ws2.append(sr)

# Style summary
for ri2, sr in enumerate(summary, start=1):
    for ci2 in range(1, 3):
        c = ws2.cell(row=ri2, column=ci2)
        c.border    = border
        c.alignment = left_al
        if sr and len(sr) > 0 and sr[0] in (
            "dokanelbanat.com — 30-Day Campaign Calendar",
            "Metric", "CONTENT TYPE", "WEEK", "PLATFORM"
        ):
            c.fill = hdr_fill
            c.font = hdr_font
        elif ri2 == 1:
            c.fill = hdr_fill
            c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=13)
        else:
            c.font = Font(name="Calibri", size=10)

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 20
ws2.row_dimensions[1].height = 28

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)

# ── Verify ────────────────────────────────────────────────────────────────────
import pathlib
p = pathlib.Path(OUTPUT_PATH)
if p.exists():
    size_kb = p.stat().st_size / 1024
    print(f"SUCCESS: {OUTPUT_PATH}")
    print(f"Size: {p.stat().st_size:,} bytes ({size_kb:.1f} KB)")
    print(f"Total rows: {len(ROWS)}")
    print(f"Content types: {type_counts}")
    print(f"Platform breakdown: {platform_counts}")
    print(f"Week breakdown: {week_counts}")
else:
    print("ERROR: File was not created!")
    sys.exit(1)

print("\nCALENDAR DONE")

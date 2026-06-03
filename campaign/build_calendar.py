import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── palette ──────────────────────────────────────────────────────────────────
COLOR_PRIMARY   = "FF009F"   # header fill
COLOR_SECONDARY = "FF2D55"
COLOR_ORANGE    = "FF6B4A"
COLOR_YELLOW    = "FFBE3D"
COLOR_GREEN     = "3DB5AA"
COLOR_LIGHT     = "FFF5FB"
WHITE           = "FFFFFF"
ALT_ROW         = "FFF0F8"   # very light pink for alternating rows

# ── output path ──────────────────────────────────────────────────────────────
OUTPUT_PATH = r"D:\claude-Projects\dokanelbanat\campaign\calendar.xlsx"
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# ── column headers ────────────────────────────────────────────────────────────
HEADERS = [
    "Day", "Date", "Week", "Theme",
    "Platform", "Content Type", "Topic (Arabic)",
    "Hook Line", "Design Brief Summary",
    "Image Prompt Summary", "File Path", "Status"
]

# ── calendar data ─────────────────────────────────────────────────────────────
# Each tuple: (day, date, week_label, theme, platform, content_type,
#              topic_arabic, hook_line, design_brief, image_prompt, file_path)
# Status is always "Ready to Publish" — added programmatically

rows = [

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 1 — BOLD & PROVOCATIVE (Days 1-7 | May 1-7)
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 1
    (1, "2026-05-01", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Carousel",
     "ماكينة الإنفاق العالمية على المرأة",
     "كل ما تحسي إنك ناقصة — في حد بيكسب منها.",
     "6-slide carousel. Slide 1: bold #ff009f headline on dark bg. Slides 2-5: data stats with icon graphics. Slide 6: dokanelbanat.com CTA. Rubik font throughout.",
     "A giant chrome machine with conveyor belt feeding shopping bags into a silhouette of a woman, neon pink on black background, editorial style.",
     "campaign/week-01-bold/posts/day-01-carousel.md",
     "Ready to Publish"),

    (1, "2026-05-01", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Single Post",
     "لماذا تشعرين دائماً بالنقص؟",
     "السؤال اللي محدش بيسأله — ليه احساس النقص ده مش بيروح أبداً؟",
     "Single image with bold Arabic question overlaid on split background: left side consumerist imagery, right side calm woman. #ff009f accent text.",
     "Split frame image: left side chaotic shopping mall in pink neon, right side serene Arab woman looking away, minimalist.",
     "campaign/week-01-bold/posts/day-01-single.md",
     "Ready to Publish"),

    # DAY 2
    (2, "2026-05-02", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Story Sequence",
     "انستجرام وصورة الجسم عند المرأة العربية",
     "5 stories. دقيقتين على انستجرام = تغيير في صورتك عن نفسك.",
     "5-story sequence. Story 1: bold question. Stories 2-4: stats with swipe-up animations. Story 5: poll sticker 'حسيتي بكده؟'. #ff009f + #ff2d55 gradient backgrounds.",
     "Phone screen mockup showing Instagram feed of filtered bodies, with a crack/glitch effect spreading across the screen, pink tones.",
     "campaign/week-01-bold/stories/day-02-story-sequence.md",
     "Ready to Publish"),

    (2, "2026-05-02", "Week 1 – Bold", "Expose the Machine",
     "Blog", "Blog Article",
     "انستجرام وأزمة صورة الجسم عند المرأة العربية",
     "كيف تحوّل انستجرام المرأة العربية إلى زبونة دائمة؟",
     "1800-word deep-dive article. Sections: research overview, psychological triggers, Arab-specific data, how algorithms exploit insecurity. Ends with dokanelbanat.com awareness framework.",
     "Hero image: Arab woman looking at phone with distorted mirror reflection showing idealized body, painterly illustration style, #ff009f dominant.",
     "campaign/blog/article-01-instagram-body-image.md",
     "Ready to Publish"),

    # DAY 3
    (3, "2026-05-03", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Infographic",
     "صناعة الجمال: $677 مليار دولار",
     "$677,000,000,000 — ده حجم الصناعة اللي بتتبنى على إحساسك بالنقص.",
     "Single tall infographic. Timeline from 1920s to 2026 showing industry growth. Icons for each decade. Bold numbers in #ffbe3d on dark bg. dokanelbanat.com watermark.",
     "Infographic poster: vertical timeline with golden dollar amounts growing, small female silhouettes at each era, dark charcoal background with pink accents.",
     "campaign/week-01-bold/infographics/day-03-infographic.md",
     "Ready to Publish"),

    (3, "2026-05-03", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Story Sequence",
     "تاريخ إعلانات المرأة ربة المنزل",
     "من 'ابشري يا ست البيت' لـ'بشرتك مش كفاية' — الإعلان دايماً عارف يضغط على الجرح.",
     "4-story sequence showing evolution of advertising targeting housewives. Vintage ad aesthetics transitioning to modern social media ads. Poll on final story.",
     "Collage of vintage Arabic housewife ads transitioning into modern beauty influencer content, sepia to neon pink color grade shift.",
     "campaign/week-01-bold/stories/day-03-story-sequence.md",
     "Ready to Publish"),

    # DAY 4
    (4, "2026-05-04", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Carousel",
     "سيكولوجية الشراء عند الإنفلوانسر",
     "مش بتشتري المنتج — بتشتري النسخة اللي إنتي عايزة تبقيها.",
     "7-slide carousel. Opens with influencer lifestyle shot. Each slide deconstructs one psychological trigger: FOMO, social proof, aspirational identity, scarcity, relatability, parasocial bond. Close with awareness CTA.",
     "Stylized illustration of a woman looking at a glowing phone; her reflection in the screen is someone different, luxury styling, neon pink palette.",
     "campaign/week-01-bold/posts/day-04-carousel.md",
     "Ready to Publish"),

    (4, "2026-05-04", "Week 1 – Bold", "Expose the Machine",
     "Blog", "Blog Article",
     "سيكولوجية الشراء: لماذا نصدق الإنفلوانسر؟",
     "العلم وراء الزر 'اشتري دلوقتي' وليه دماغك مش قادرة تقاومه.",
     "2000-word article covering parasocial relationships, aspirational identity, neurological dopamine triggers, FOMO mechanics in influencer marketing. Arab market examples included.",
     "Hero: close-up of an Arab woman's eye reflected in a phone screen showing a glowing 'buy now' button, cinematic lighting, pink and gold tones.",
     "campaign/blog/article-02-influencer-psychology.md",
     "Ready to Publish"),

    # DAY 5
    (5, "2026-05-05", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Thread",
     "نظرية المقارنة الاجتماعية وتأثيرها على المرأة",
     "ليه ما تقدريش تبطلي تقارني نفسك بغيرك — وده مش عيب فيكي.",
     "8-post thread. Each post covers: Festinger's original theory, upward vs downward comparison, social media amplification, body image link, self-esteem impact, Arab cultural layer, breaking the cycle, dokanelbanat.com community as counter-space.",
     "Series of minimal cards: each features a quote or stat on clean white/pink background with subtle geometric patterns. Thread visual identity consistent.",
     "campaign/week-01-bold/posts/day-05-thread.md",
     "Ready to Publish"),

    (5, "2026-05-05", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Single Post",
     "المقارنة الاجتماعية في السياق العربي",
     "المجتمع بيعلمنا نقارن — والسوشيال ميديا بتتاجر فيها.",
     "Single bold image with Arabic text about social comparison in Arab cultural context. Stats overlay. Warm #ff6b4a tone with white text.",
     "Two Arab women side by side, one showing phone, illustrated style, warm orange tones, comparing lifestyle imagery, editorial feel.",
     "campaign/week-01-bold/posts/day-05-single.md",
     "Ready to Publish"),

    # DAY 6
    (6, "2026-05-06", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Story Sequence",
     "إنفاق المرأة العربية على منتجات التجميل",
     "إيه اللي بتنفقيه فعلاً — وعلى إيه بالظبط؟",
     "5-story sequence with interactive elements. Story 1: shocking spend stat. Stories 2-3: breakdown by category. Story 4: quiz 'كام صرفتي الشهر ده على التجميل؟'. Story 5: awareness reframe.",
     "Animated-style story frames: coins and beauty products floating, Arabic text bold white on #ff009f, calculator emoji integration, clean design.",
     "campaign/week-01-bold/stories/day-06-story-sequence.md",
     "Ready to Publish"),

    (6, "2026-05-06", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Infographic",
     "أين تذهب أموال المرأة العربية؟",
     "خريطة الإنفاق — من جيبك لجيب مين؟",
     "Horizontal spending breakdown infographic. Pie chart plus category icons: skincare, makeup, fashion, hair, diet products. Each slice in brand palette colors. Arabic labels.",
     "Overhead flat-lay of beauty products arranged as a pie chart, each section a different color matching brand palette, white marble background.",
     "campaign/week-01-bold/infographics/day-06-infographic.md",
     "Ready to Publish"),

    # DAY 7
    (7, "2026-05-07", "Week 1 – Bold", "Expose the Machine",
     "Instagram", "Carousel",
     "ملخص الأسبوع الأول: الماكينة كشفت",
     "أسبوع كامل من الحقايق — إيه اللي صدمك أكتر؟",
     "6-slide week-recap carousel. Slide 1: Week 1 title card. Slides 2-5: One bold takeaway per day. Slide 6: tease Week 2 'الأسبوع الجاي: الأرقام والبيانات'. CTA to follow for Week 2.",
     "Mosaic recap design: mini thumbnails of each Week 1 topic in grid, bold Week 1 summary headline, #ff009f dominant, Rubik bold typography.",
     "campaign/week-01-bold/posts/day-07-carousel.md",
     "Ready to Publish"),

    (7, "2026-05-07", "Week 1 – Bold", "Expose the Machine",
     "Facebook", "Thread",
     "أسئلة الأسبوع الأول من المجتمع",
     "إنتو سألتوا — دي إجاباتنا. الحوار الحقيقي بيبدأ هنا.",
     "Facebook thread post collecting and answering top 5 community questions from Week 1. Each answer is a separate comment thread. Engages existing audience.",
     "Community conversation bubble design: multiple speech bubbles in brand colors with Arabic question marks and response indicators.",
     "campaign/week-01-bold/posts/day-07-thread.md",
     "Ready to Publish"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 2 — EDUCATE & DATA (Days 8-14 | May 8-14)
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 8
    (8, "2026-05-08", "Week 2 – Educate", "Research & Data",
     "Instagram", "Carousel",
     "أحدث أبحاث تأثير السوشيال ميديا على المرأة",
     "مش رأي — ده بحث علمي. الأرقام بتتكلم.",
     "8-slide data carousel. Each slide = one research finding with source citation, clean data visualization, academic-meets-accessible design. #3db5aa (green) accent for credibility.",
     "Clean research poster aesthetic: white cards with data charts, green accent bars, Arabic citations, scientific feel with warm brand colors.",
     "campaign/week-02-educate/posts/day-08-carousel.md",
     "Ready to Publish"),

    (8, "2026-05-08", "Week 2 – Educate", "Research & Data",
     "Blog", "Blog Article",
     "الأبحاث العلمية وتأثير السوشيال ميديا على صورة المرأة",
     "ما تقوله الأبحاث عن السوشيال ميديا وصورة المرأة — والأرقام مفاجئة.",
     "2200-word research synthesis article. Covers 5 key studies: body image, self-esteem, purchase behavior, social comparison, addiction loops. Full citations. Accessible Arabic language.",
     "Hero: academic research papers overlaid with social media icons, illustrated style, green and pink color palette, bookish but modern aesthetic.",
     "campaign/blog/article-03-research-social-media.md",
     "Ready to Publish"),

    # DAY 9
    (9, "2026-05-09", "Week 2 – Educate", "Research & Data",
     "Instagram", "Thread",
     "آليات الإنفلوانسر: كيف يعمل النظام؟",
     "من العقد للمحتوى للضغط النفسي — رحلة المنتج من المصنع لعقلك.",
     "7-post educational thread. Posts cover: brand-influencer contracts, disclosure laws vs. reality, content authenticity illusion, FTC vs. Arab market regulations, micro vs. macro influencer ROI, the 'relatability' construction, audience data harvesting.",
     "Thread card series: flowchart-style with arrows showing influencer ecosystem, brand boxes, money flow arrows, psychological impact icons. Clean infographic cards.",
     "campaign/week-02-educate/posts/day-09-thread.md",
     "Ready to Publish"),

    (9, "2026-05-09", "Week 2 – Educate", "Research & Data",
     "Facebook", "Story Sequence",
     "كيف يتم اختيار الإنفلوانسر للترويج لمنتجك؟",
     "خلف الكواليس — العملية اللي ما بتشوفيهاش.",
     "4-story sequence exposing influencer-brand selection process. Behind-the-scenes infographic style. Swipe-to-reveal mechanic concept. Platform-native interactive stickers.",
     "Behind-the-scenes documentary style: contract papers, brand briefs, engagement rate spreadsheets layered with Instagram grid mockups, moody blue-pink lighting.",
     "campaign/week-02-educate/stories/day-09-story-sequence.md",
     "Ready to Publish"),

    # DAY 10
    (10, "2026-05-10", "Week 2 – Educate", "Research & Data",
      "Instagram", "Infographic",
      "أنماط الإعلان التاريخية التي استهدفت المرأة",
      "من الخمسينيات للألفية — الإعلان دايماً عارف ضعفك.",
      "Timeline infographic. Decade by decade: 1950s housewife ads, 1970s liberation paradox, 1990s heroin chic, 2000s celebrity, 2010s natural beauty fake-out, 2020s body positivity used as marketing. Vintage to modern visual style.",
      "Vertical vintage-to-modern timeline collage: each decade has a mini ad mockup, transitioning from sepia newspaper style to neon digital screen aesthetic.",
      "campaign/week-02-educate/infographics/day-10-infographic.md",
      "Ready to Publish"),

    (10, "2026-05-10", "Week 2 – Educate", "Research & Data",
      "Facebook", "Single Post",
      "الإعلان التاريخي وتأثيره على المرأة العربية",
      "إزاي السوق تعلم المرأة العربية تكره جسمها — وإمتى بدأ ده؟",
      "Single educational post with key historical insight, long-form Facebook caption format, detailed explanation of advertising's role in Arab women's body image history.",
      "Sepia-toned vintage Arabic magazine ad reimagined in modern graphic style, split with contemporary social media equivalent, contrast effect.",
      "campaign/week-02-educate/posts/day-10-single.md",
      "Ready to Publish"),

    # DAY 11
    (11, "2026-05-11", "Week 2 – Educate", "Research & Data",
      "Instagram", "Carousel",
      "سوق الجمال في منطقة MENA: 46 مليار دولار",
      "46 مليار دولار — ومصر والسعودية في القلب.",
      "6-slide market data carousel. Slides: MENA market size, Egypt breakdown, Saudi breakdown, fastest growing segments, key players, what this means for Arab women consumers. Data-heavy but visually engaging.",
      "Market data visualization: MENA map with glowing data points, bar charts in brand colors, financial report aesthetic with Arabic typography overlay.",
      "campaign/week-02-educate/posts/day-11-carousel.md",
      "Ready to Publish"),

    (11, "2026-05-11", "Week 2 – Educate", "Research & Data",
      "Blog", "Blog Article",
      "سوق الجمال في الشرق الأوسط: 46 مليار دولار وأنتِ في المنتصف",
      "من يستفيد فعلاً من سوق الجمال في الشرق الأوسط؟",
      "2000-word market analysis article. MENA beauty market structure, who controls it, how Arab women consumers are positioned, alternative conscious consumption models. dokanelbanat.com as community counter-model.",
      "Hero: stylized MENA map made of beauty product icons, glowing green and pink, data overlay aesthetic, editorial magazine quality.",
      "campaign/blog/article-04-mena-beauty-market.md",
      "Ready to Publish"),

    # DAY 12
    (12, "2026-05-12", "Week 2 – Educate", "Research & Data",
      "Instagram", "Story Sequence",
      "إحصائيات تشوه صورة الجسم Body Dysmorphia",
      "الأرقام اللي محتاجة تعرفيها — ومش الصحافة بتتكلم فيها.",
      "5-story educational sequence on body dysmorphia statistics. Story 1: shocking global stat. Story 2: Arab-specific data. Story 3: social media correlation. Story 4: age of onset data. Story 5: resources + dokanelbanat.com community CTA.",
      "Clinical-meets-human design: soft watercolor illustration of a woman looking in a broken mirror, statistics overlaid as bold text, green and pink tones.",
      "campaign/week-02-educate/stories/day-12-story-sequence.md",
      "Ready to Publish"),

    (12, "2026-05-12", "Week 2 – Educate", "Research & Data",
      "Facebook", "Carousel",
      "تشوه صورة الجسم: أرقام وحقائق",
      "ما بين الوسواس والتسوق — الصلة الخفية اللي العلم أثبتها.",
      "5-slide carousel connecting body dysmorphia statistics to compulsive purchasing behavior. Cites peer-reviewed research. Sensitive but direct tone. Professional mental health messaging.",
      "Mosaic of statistics: gentle illustrated facial features with data numbers embedded, pastel base with #ff009f accents, informational yet compassionate visual style.",
      "campaign/week-02-educate/posts/day-12-carousel.md",
      "Ready to Publish"),

    # DAY 13
    (13, "2026-05-13", "Week 2 – Educate", "Research & Data",
      "Instagram", "Infographic",
      "بحث إدمان السوشيال ميديا: الخوارزمية ضدك",
      "الخوارزمية مش بتوصيلك محتوى — بتوصيلك إحساس.",
      "Tall infographic explaining social media addiction research: dopamine loop, variable reward schedule, endless scroll design, comparison triggers, attention economy. Each element with research citation.",
      "Brain illustration with social media platforms as neural pathways, dopamine molecules as pink bubbles, dark background with neon pink and green data elements.",
      "campaign/week-02-educate/infographics/day-13-infographic.md",
      "Ready to Publish"),

    (13, "2026-05-13", "Week 2 – Educate", "Research & Data",
      "Facebook", "Thread",
      "إدمان السوشيال ميديا: هل أنتِ مدمنة؟",
      "7 أسئلة تحددي بيهم علاقتك الحقيقية مع السوشيال ميديا.",
      "Facebook thread with 7 self-assessment questions about social media addiction patterns. Each question has brief explanation. Ends with free resource from dokanelbanat.com.",
      "Thread visual: question card design in bold Arabic numerals, #ff009f circles, clean white background, assessment worksheet aesthetic.",
      "campaign/week-02-educate/posts/day-13-thread.md",
      "Ready to Publish"),

    # DAY 14
    (14, "2026-05-14", "Week 2 – Educate", "Research & Data",
      "Instagram", "Carousel",
      "دراسة السعودية: صورة الجسم والشبكات الاجتماعية",
      "دراسة سعودية تكشف: الفتيات العربيات أكثر تأثراً — وده السبب.",
      "7-slide carousel presenting Saudi Arabia body image study findings. Contextualizes Arab cultural specificities: modesty culture paradox, family comparison dynamics, wedding culture pressure, whitening product demand.",
      "Research presentation style: slide cards with Saudi cultural imagery (modest clothing, family settings) intersected with data charts, warm tan and pink color palette.",
      "campaign/week-02-educate/posts/day-14-carousel.md",
      "Ready to Publish"),

    (14, "2026-05-14", "Week 2 – Educate", "Research & Data",
      "Blog", "Blog Article",
      "الدراسة السعودية: كيف تؤثر وسائل التواصل على صورة الجسم في العالم العربي",
      "ما كشفته الدراسات العربية عن صورة الجسم — والأرقام مقلقة.",
      "2100-word article summarizing key Arab-region studies on body image and social media, focusing on Saudi research, Egypt comparisons, cultural specificities of beauty pressure in Arab societies.",
      "Hero: illustrated map of Arab world with data visualization overlay, female silhouettes embedded, research journal aesthetic, green and pink gradient.",
      "campaign/blog/article-05-saudi-body-image-study.md",
      "Ready to Publish"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 3 — SHIFT (Days 15-21 | May 15-21)
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 15
    (15, "2026-05-15", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "بداية التحول: اخترنا الوعي",
      "الوعي مش قرار — هو عادة. وبتبدأ هنا.",
      "6-slide pivot carousel. Shift in visual tone: lighter backgrounds, #fff5fb dominant. Introduces mindset shift concept. Contrasts week 1-2 darkness with new possibility. Announces the 4 pillars framework.",
      "Sunrise aesthetic: woman stretching outdoors in morning light, brand color sunrise gradient, clean and airy feel, Rubik bold Arabic text overlay.",
      "campaign/week-03-shift/posts/day-15-carousel.md",
      "Ready to Publish"),

    (15, "2026-05-15", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Single Post",
      "التحول نحو الاستهلاك الواعي",
      "إزاي تبدأي رحلة الوعي من غير ما تحسي بذنب على الماضي.",
      "Single reflective post with warm tone. Long-form Facebook caption about beginning a conscious consumption journey. Non-judgmental, empowering language. Includes 3 first steps.",
      "Warm sunrise flat lay: journal, simple skincare products, morning coffee, natural light photography aesthetic, soft pink and cream tones.",
      "campaign/week-03-shift/posts/day-15-single.md",
      "Ready to Publish"),

    # DAY 16
    (16, "2026-05-16", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Story Sequence",
      "الحد الأدنى من الجمال: فلسفة المينيماليزم",
      "أقل = أكتر. فلسفة المينيماليزم في الجمال والحياة.",
      "5-story sequence on beauty minimalism. Story 1: define minimalism in beauty context. Stories 2-3: capsule skincare concept, capsule wardrobe concept. Story 4: 5 questions to audit your current routine. Story 5: CTA to dokanelbanat.com community.",
      "Minimal aesthetic stories: white and cream backgrounds, simple flat-lay product groupings, 'less is more' typography, clean geometric framing.",
      "campaign/week-03-shift/stories/day-16-story-sequence.md",
      "Ready to Publish"),

    (16, "2026-05-16", "Week 3 – Shift", "Alternative Mindset",
      "Blog", "Blog Article",
      "المينيماليزم والجمال: كيف تبنين روتيناً واعياً بأقل المنتجات",
      "الروتين الواعي مش محتاج 20 منتج — 5 كفاية.",
      "1900-word guide to beauty minimalism. Covers: capsule skincare philosophy, ingredient awareness, multi-use products, slow beauty movement, Arab skin types guide, dokanelbanat.com curated recommendations.",
      "Hero: flat-lay of 5 minimal skincare products on marble, soft natural lighting, cream and pink tones, clean editorial photography style.",
      "campaign/blog/article-06-minimalism-beauty.md",
      "Ready to Publish"),

    # DAY 17
    (17, "2026-05-17", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "الروتين اليومي الواعي: خطوة بخطوة",
      "يومك من غير ما السوشيال ميديا تحكمه — روتين يومي واعي.",
      "7-slide carousel presenting a complete conscious daily routine. Morning, afternoon, evening. Each time block has: awareness check, intentional choice, gratitude element. Practical and aspirational.",
      "Day-planner aesthetic: time blocks illustrated with warm icons, #ffbe3d morning, #3db5aa afternoon, #ff009f evening accents, clean calendar-style layout.",
      "campaign/week-03-shift/posts/day-17-carousel.md",
      "Ready to Publish"),

    (17, "2026-05-17", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Story Sequence",
      "كيف تبدئين يومك بدون سوشيال ميديا؟",
      "أول 30 دقيقة في يومك — اللي بتقرري فيها مزاجك كله.",
      "4-story practical guide. Story 1: why first 30 minutes matter. Stories 2-3: alternative morning routine steps. Story 4: 7-day challenge invitation linked to dokanelbanat.com community.",
      "Morning routine photography style: sunrise, journal, no-phone visual, woman meditating, warm cream and gold tones, peaceful aesthetic.",
      "campaign/week-03-shift/stories/day-17-story-sequence.md",
      "Ready to Publish"),

    # DAY 18
    (18, "2026-05-18", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Thread",
      "الاكتفاء الذاتي: المرأة اللي تكتفي بنفسها",
      "الاكتفاء مش عزلة — هو قوة. الفرق بين الاثنين.",
      "8-post thread on self-sufficiency for Arab women. Posts: define self-sufficiency vs. isolation, financial self-sufficiency, emotional self-sufficiency, knowledge self-sufficiency, skill-building, community vs. dependency, Arab cultural context, dokanelbanat.com as resource.",
      "Thread card series: strong clean design, each card has Arabic keyword bold center, supporting stat or quote below, #ff009f and #3db5aa alternating accents.",
      "campaign/week-03-shift/posts/day-18-thread.md",
      "Ready to Publish"),

    (18, "2026-05-18", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Infographic",
      "أربعة أركان المرأة الواعية",
      "4 أركان — 1 هوية. اكتشفي النموذج اللي بيبنيكي.",
      "Square infographic (4-quadrant layout). Each quadrant = one pillar: Self-Sufficiency الاكتفاء الذاتي / Conscious Awareness الوعي الاستهلاكي / Routine الروتين المتعمد / Empowering Community المجتمع الداعم. Brand colors per quadrant.",
      "Four-quadrant illustration: each quadrant a different brand color (#ff009f, #ff6b4a, #ffbe3d, #3db5aa), icon and Arabic label centered, clean geometric design.",
      "campaign/week-03-shift/infographics/day-18-infographic.md",
      "Ready to Publish"),

    # DAY 19
    (19, "2026-05-19", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "فلسفة الجمال النظيف: Clean Beauty عربي",
      "Clean Beauty مش ترند — هو حق. المرأة العربية والمكونات النظيفة.",
      "6-slide carousel. Slide 1: clean beauty definition in Arab context. Slides 2-4: common harmful ingredients to avoid (Arabic names), safe alternatives, Arab-specific skin concerns. Slide 5: reading labels guide. Slide 6: dokanelbanat.com clean picks.",
      "Clean product flat-lay series: white marble, fresh botanicals, Arabic ingredient labels visible, #3db5aa green accent for 'clean', scientific but beautiful aesthetic.",
      "campaign/week-03-shift/posts/day-19-carousel.md",
      "Ready to Publish"),

    (19, "2026-05-19", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Thread",
      "كيف تقرئين مكونات المنتج بالعربي؟",
      "قبل ما تشتري أي منتج — اقرأي ده الأول.",
      "Facebook thread: step-by-step guide to reading product ingredient labels. 6 comments with specific guidance. Includes common Arabic/INCI name translations. Actionable and educational.",
      "Thread cards: product ingredient list mockup with highlighted warning ingredients, magnifying glass illustration, clean educational aesthetic, green tick/red cross system.",
      "campaign/week-03-shift/posts/day-19-thread.md",
      "Ready to Publish"),

    # DAY 20
    (20, "2026-05-20", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Story Sequence",
      "معايير الجمال الحقيقية: من قال إنها صح؟",
      "من قرر إن الشعر الناعم أحسن؟ ومن قرر إن البشرة الفاتحة أجمل؟",
      "5-story sequence deconstructing beauty standards. Story 1: who decides beauty standards? Story 2: colonial history of Arab beauty standards. Story 3: diversity of natural Arab beauty. Story 4: media's role in standard-setting. Story 5: empowerment reframe.",
      "Story series: diverse Arab women portraits (illustrated), each with a beauty standard myth as strikethrough text, vibrant portrait colors, empowering Arabic captions.",
      "campaign/week-03-shift/stories/day-20-story-sequence.md",
      "Ready to Publish"),

    (20, "2026-05-20", "Week 3 – Shift", "Alternative Mindset",
      "Blog", "Blog Article",
      "معايير الجمال العربية: من صنعها ولماذا؟",
      "الجمال العربي تحت المجهر — من وضع المعايير ولمصلحة من؟",
      "2000-word article examining Arab beauty standards historically and commercially. Covers colonialism's impact, media manufacturing of standards, the lightening cream industry, natural Arab diversity celebration.",
      "Hero: illustrated collage of diverse Arab women in natural settings, warm earth tones, deconstructed beauty product imagery in background, empowering editorial style.",
      "campaign/blog/article-07-arab-beauty-standards.md",
      "Ready to Publish"),

    # DAY 21
    (21, "2026-05-21", "Week 3 – Shift", "Alternative Mindset",
      "Instagram", "Carousel",
      "ملخص الأسبوع الثالث: التحول اكتمل",
      "3 أسابيع — والتحول بدأ. إيه اللي تغير في طريقة تفكيرك؟",
      "6-slide week recap. Summarizes shift week insights: minimalism, 4 pillars, clean beauty, real standards, daily routine. Introduces Week 4 Empower theme. Community engagement CTA.",
      "Mosaic of Week 3 content mini-previews, lighter airy aesthetic reflecting the shift, brand colors in lighter tones, empowering Arabic headlines, Week 4 teaser slide.",
      "campaign/week-03-shift/posts/day-21-carousel.md",
      "Ready to Publish"),

    (21, "2026-05-21", "Week 3 – Shift", "Alternative Mindset",
      "Facebook", "Infographic",
      "مقارنة: قبل الوعي وبعده",
      "جدول المقارنة — حياتك قبل الوعي الاستهلاكي وبعده.",
      "Before/After comparison infographic. Two columns: 'قبل الوعي' (dark/chaotic visuals) vs 'بعد الوعي' (clean/calm visuals). 8 comparison points across spending, self-image, community, routine, knowledge.",
      "Split comparison table design: left column dark chaotic, right column clean calm, same woman illustrated in both states, brand transformation visual.",
      "campaign/week-03-shift/infographics/day-21-infographic.md",
      "Ready to Publish"),

    # ══════════════════════════════════════════════════════════════════════════
    # WEEK 4 — EMPOWER (Days 22-30 | May 22-30)
    # ══════════════════════════════════════════════════════════════════════════

    # DAY 22
    (22, "2026-05-22", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "أنتِ مكتملة: البيان الأول",
      "مش محتاجة أي منتج علشان تبقي مكتملة. أنتِ مكتملة دلوقتي.",
      "7-slide empowerment manifesto carousel. Slide 1: bold 'أنتِ مكتملة' statement. Slides 2-6: each pillar of completeness (inner/outer/community/knowledge/purpose). Slide 7: dokanelbanat.com as home for complete women.",
      "Power poster aesthetic: full-frame Arabic bold text, each slide a different brand gradient, strong typographic design, no product imagery — only words and color.",
      "campaign/week-04-empower/posts/day-22-carousel.md",
      "Ready to Publish"),

    (22, "2026-05-22", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Single Post",
      "رسالة dokanelbanat.com: لماذا أنشأنا هذه المنصة",
      "السبب الحقيقي وراء dokanelbanat.com — مش متجر. مجتمع.",
      "Founder's voice post. Long-form Facebook narrative about why dokanelbanat.com was created, the problem it solves, the community it's building. Personal, authentic, mission-driven.",
      "Warm editorial portrait style: woman at desk with journal and laptop, dokanelbanat.com branding subtle in background, candid and real, #fff5fb tones.",
      "campaign/week-04-empower/posts/day-22-single.md",
      "Ready to Publish"),

    # DAY 23
    (23, "2026-05-23", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Story Sequence",
      "قصص النساء الرائدات من المجتمع",
      "مش بنحكي عنهم — بنحكي معاهم. قصص حقيقية من مجتمعنا.",
      "5-story community spotlight sequence. Each story features a real community member (anonymized or with permission): her challenge, her shift, her win. Story 5: invite submissions for next spotlight.",
      "Portrait spotlight design: soft vignette around illustrated portrait, quote pull in Arabic, #ff009f accent frame, community member name in Rubik bold.",
      "campaign/week-04-empower/stories/day-23-story-sequence.md",
      "Ready to Publish"),

    (23, "2026-05-23", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "نساء رائدات: قصص من مجتمع dokanelbanat.com",
      "نساء عربيات اخترن الوعي — قصصهم بتغير مجتمعات.",
      "1800-word community spotlight article. 4-5 stories of Arab women entrepreneurs and conscious consumers from the community. Each story: background, challenge, mindset shift, what changed. dokanelbanat.com as catalyst.",
      "Hero: illustrated group portrait of diverse Arab women in professional settings, warm community feel, editorial quality, brand color accents.",
      "campaign/blog/article-08-community-women-stories.md",
      "Ready to Publish"),

    # DAY 24
    (24, "2026-05-24", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Infographic",
      "النساء رائدات الأعمال في العالم العربي: أرقام الإلهام",
      "المرأة العربية رائدة أعمال — والأرقام بتقول حاجة تانية.",
      "Infographic on Arab women entrepreneurs: growth stats, key sectors, Egypt and Gulf data, barriers vs. opportunities, dokanelbanat.com as support ecosystem. Empowering data presentation.",
      "Bold entrepreneurship infographic: rising arrow charts, female silhouette with business elements, #ff009f and #ffbe3d dominant, Arabic stat callouts, confident visual tone.",
      "campaign/week-04-empower/infographics/day-24-infographic.md",
      "Ready to Publish"),

    (24, "2026-05-24", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Carousel",
      "كيف تبدئين مشروعك الصغير من الصفر؟",
      "من الفكرة للبيعة الأولى — دليل عملي للمرأة العربية.",
      "6-slide practical carousel for aspiring women entrepreneurs. Steps: idea validation, minimal viable product, first sales channel, pricing strategy, building audience, where dokanelbanat.com fits in the journey.",
      "Step-by-step guide cards: numbered steps with icon illustrations, warm orange (#ff6b4a) dominant, professional yet approachable, small business imagery.",
      "campaign/week-04-empower/posts/day-24-carousel.md",
      "Ready to Publish"),

    # DAY 25
    (25, "2026-05-25", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Thread",
      "منظومة dokanelbanat.com: كل ما تحتاجينه في مكان واحد",
      "واحدة سألتني: إيه اللي عنده dokanelbanat.com فعلاً؟ — اتفضلي.",
      "8-post thread walking through the dokanelbanat.com ecosystem: awareness content, educational resources, community forum, product curation, entrepreneur support, expert network, events, newsletter. Each post = one pillar.",
      "Ecosystem map thread cards: hub-and-spoke diagram, dokanelbanat.com center, each feature as a spoke, consistent card design, brand colors per section.",
      "campaign/week-04-empower/posts/day-25-thread.md",
      "Ready to Publish"),

    (25, "2026-05-25", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "منظومة dokanelbanat.com: كيف تستفيدين من كل خدمة؟",
      "كل ما قدمته dokanelbanat.com — وكيف تستخدمينه لصالحك.",
      "2000-word complete ecosystem guide. Each section covers one dokanelbanat.com offering with specific use cases, how to access, and real benefit for Arab women. SEO-optimized for brand awareness.",
      "Hero: dokanelbanat.com website mockup surrounded by community icons, educational elements, and product imagery, brand-accurate color palette, editorial quality.",
      "campaign/blog/article-09-dokanelbanat-ecosystem.md",
      "Ready to Publish"),

    # DAY 26
    (26, "2026-05-26", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "بناء مجتمع مخلص: الفرق بين الجمهور والمجتمع",
      "مليون فولوور مش مجتمع — وألف امرأة واعية أقوى منهم.",
      "7-slide carousel on community vs. audience building. Slide 1: the distinction. Slides 2-5: what makes a real community (shared values, two-way dialogue, mutual support, growth). Slide 6: dokanelbanat.com community principles. Slide 7: join CTA.",
      "Community visualization carousel: concentric circles showing depth over breadth, warm human illustration style, conversations and connections depicted, #3db5aa green for community slides.",
      "campaign/week-04-empower/posts/day-26-carousel.md",
      "Ready to Publish"),

    (26, "2026-05-26", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Story Sequence",
      "كيف تنضمين لمجتمع dokanelbanat.com؟",
      "3 خطوات بس — وتبقي جزء من أقوى مجتمع نساء عربيات.",
      "3-story onboarding sequence for community joining. Story 1: what you get as a member. Story 2: how to join (step-by-step). Story 3: first challenge/action to take after joining. Clear CTA throughout.",
      "Onboarding welcome design: open door visual metaphor, warm welcome gradient, step-by-step clean icons, community member avatar cluster, inviting brand colors.",
      "campaign/week-04-empower/stories/day-26-story-sequence.md",
      "Ready to Publish"),

    # DAY 27
    (27, "2026-05-27", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Story Sequence",
      "الاستهلاك الواعي في 30 يوم: ماذا تعلمنا؟",
      "30 يوم من الوعي — والرحلة مش خلصت، هي بس بدأت.",
      "5-story campaign retrospective. Story 1: campaign journey overview. Story 2: most impactful Week 1 reveal. Story 3: most shared Week 2 data. Story 4: most resonant Week 3 shift. Story 5: Week 4 empowerment summary + what's next.",
      "Retrospective story design: mini-film-strip layout, each story showing campaign highlights, nostalgic-meets-forward-looking aesthetic, brand gradient evolution from dark to light.",
      "campaign/week-04-empower/stories/day-27-story-sequence.md",
      "Ready to Publish"),

    (27, "2026-05-27", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Thread",
      "30 يوم من الوعي: الدروس الكبيرة",
      "كل يوم علمنا حاجة — دي أهم 10 دروس من 30 يوم.",
      "Facebook thread: 10 key learnings from the 30-day campaign. Each comment = one lesson with brief explanation, the community data/engagement that validated it, and the forward implication.",
      "Campaign learning cards: numbered lessons in a clean list format, #ff009f bullet accents, summary design that could work as a shareable screenshot.",
      "campaign/week-04-empower/posts/day-27-thread.md",
      "Ready to Publish"),

    # DAY 28
    (28, "2026-05-28", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "بناء البراند الشخصي للمرأة العربية",
      "براندك الشخصي مش لوجو ولا كولر — هو قيمتك في السوق.",
      "6-slide personal branding carousel for Arab women. Covers: what personal brand means for Arab women, values-based positioning, authentic storytelling, LinkedIn for Arab professional women, dokanelbanat.com community as brand amplifier.",
      "Personal brand builder carousel: clean professional design, woman-at-work imagery, brand identity elements (colors/fonts/voice) illustrated, #ff009f accent, empowering business aesthetic.",
      "campaign/week-04-empower/posts/day-28-carousel.md",
      "Ready to Publish"),

    (28, "2026-05-28", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "البراند الشخصي للمرأة العربية: دليل البداية",
      "كيف تبنين براند شخصي أصيل كامرأة عربية في 2026؟",
      "2100-word personal branding guide for Arab women. Covers: discovering your unique positioning, building authentic online presence, leveraging community, specific LinkedIn and social strategies for Arab professional women.",
      "Hero: Arab professional woman at whiteboard with brand strategy elements, confident pose, warm office setting, brand colors in environment design.",
      "campaign/blog/article-10-personal-brand-arab-women.md",
      "Ready to Publish"),

    # DAY 29
    (29, "2026-05-29", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Thread",
      "الاستهلاك الواعي: دليلك العملي للأبد",
      "مش حملة — طريقة حياة. دليل الاستهلاك الواعي اللي يفضل معاكي.",
      "8-post evergreen thread. Complete conscious consumption framework: monthly spending audit, social media curation, product ingredient awareness, community validation before purchase, supporting women-owned brands, gratitude practice, quarterly mindset check, annual brand detox.",
      "Evergreen guide thread: clean handbook aesthetic, each post a 'chapter' card, #3db5aa green for life-long-use feeling, practical icons, Arabic chapter numbers bold.",
      "campaign/week-04-empower/posts/day-29-thread.md",
      "Ready to Publish"),

    (29, "2026-05-29", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Infographic",
      "خريطة الاستهلاك الواعي للمرأة العربية",
      "كل قرار شراء بيمر بـ7 خطوات — هل بتمشي عليهم؟",
      "Decision-making flowchart infographic for conscious consumption. 7-step purchase decision tree: Do I need it? Research step. Community check. Values alignment. Budget check. Environmental impact. Buy or wait. Shareable poster format.",
      "Flowchart poster: clean decision tree with yes/no branches, Arabic text throughout, color-coded branches (green = proceed, #ff009f = pause, orange = research more), clean white background.",
      "campaign/week-04-empower/infographics/day-29-infographic.md",
      "Ready to Publish"),

    # DAY 30
    (30, "2026-05-30", "Week 4 – Empower", "You Are Complete",
      "Instagram", "Carousel",
      "يوم 30: أنتِ مكتملة — البيان النهائي",
      "30 يوم، ومجتمع كامل. أنتِ مكتملة وجاهزة لتغيير الوعي.",
      "8-slide campaign finale carousel. Slide 1: bold campaign end statement. Slides 2-4: recap of 3 phase journey. Slide 5: community achievement stats. Slide 6: 4 pillars visual. Slide 7: what comes next for dokanelbanat.com. Slide 8: manifesto closing.",
      "Grand finale carousel: most impactful design of the campaign. Full-bleed brand colors, gold (#ffbe3d) celebration accents, community montage slide, bold closing manifesto typography.",
      "campaign/week-04-empower/posts/day-30-carousel.md",
      "Ready to Publish"),

    (30, "2026-05-30", "Week 4 – Empower", "You Are Complete",
      "Blog", "Blog Article",
      "30 يوم من الوعي: الرسالة الختامية لمجتمع dokanelbanat.com",
      "الرسالة الأخيرة — وبداية كل شيء.",
      "2500-word campaign closing article. Full journey narrative: what we exposed, what the data showed, the mindset shift we made together, the community we built, what dokanelbanat.com commits to going forward. Manifesto-quality writing.",
      "Hero: wide community illustration — diverse Arab women in circle, sunrise behind, empowering atmosphere, all brand colors harmonized, cinematic quality.",
      "campaign/blog/article-11-campaign-closing-manifesto.md",
      "Ready to Publish"),

    (30, "2026-05-30", "Week 4 – Empower", "You Are Complete",
      "Facebook", "Story Sequence",
      "شكراً — وما التالي مع dokanelbanat.com",
      "الحملة خلصت — المجتمع لأ. إيه اللي بيجي بعد كده؟",
      "4-story closing sequence. Story 1: thank you message with campaign stats. Story 2: what's coming next from dokanelbanat.com. Story 3: how to stay connected (newsletter, community, social). Story 4: share your transformation CTA.",
      "Closing celebration story design: confetti-style with brand color particles, warm gratitude aesthetic, forward-looking imagery, newsletter signup CTA card.",
      "campaign/week-04-empower/stories/day-30-story-sequence.md",
      "Ready to Publish"),
]

# ── build workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "30-Day Calendar"

# ── styles ────────────────────────────────────────────────────────────────────
header_fill   = PatternFill("solid", fgColor=COLOR_PRIMARY)
header_font   = Font(name="Calibri", bold=True, color=WHITE, size=11)
alt_fill      = PatternFill("solid", fgColor=ALT_ROW)
normal_fill   = PatternFill("solid", fgColor=WHITE)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin_side     = Side(style="thin", color="DDDDDD")
thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# week accent fills
week_fills = {
    "Week 1 – Bold":    PatternFill("solid", fgColor="FFE0F0"),
    "Week 2 – Educate": PatternFill("solid", fgColor="E0F8FF"),
    "Week 3 – Shift":   PatternFill("solid", fgColor="E8FFE0"),
    "Week 4 – Empower": PatternFill("solid", fgColor="FFF8E0"),
}

# ── header row ────────────────────────────────────────────────────────────────
ws.append(HEADERS)
for col_idx, _ in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[1].height = 30

# ── data rows ─────────────────────────────────────────────────────────────────
for row_idx, r in enumerate(rows, start=2):
    # append full row including Status
    ws.append(list(r) + ["Ready to Publish"])

    week_label = r[2]
    is_alt = (row_idx % 2 == 0)
    row_base_fill = week_fills.get(week_label, normal_fill)

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        # apply week tint; alternate rows slightly darker via alt_fill overlay
        if is_alt:
            cell.fill = alt_fill
        else:
            cell.fill = row_base_fill

        # alignment per column
        if col_idx in (1, 2, 3, 6, 12):   # Day, Date, Week, Content Type, Status
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    ws.row_dimensions[row_idx].height = 55

# ── freeze header row ─────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── auto-size columns ─────────────────────────────────────────────────────────
col_min_widths = {
    1: 6,   # Day
    2: 13,  # Date
    3: 16,  # Week
    4: 20,  # Theme
    5: 12,  # Platform
    6: 18,  # Content Type
    7: 35,  # Topic (Arabic)
    8: 40,  # Hook Line
    9: 55,  # Design Brief Summary
    10: 50, # Image Prompt Summary
    11: 48, # File Path
    12: 18, # Status
}

for col_idx, col_cells in enumerate(ws.columns, start=1):
    max_len = 0
    for cell in col_cells:
        if cell.value:
            # use first line only for width calc
            line = str(cell.value).split("\n")[0]
            max_len = max(max_len, len(line))
    width = min(max(max_len + 3, col_min_widths.get(col_idx, 12)), 70)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── add summary stats sheet ───────────────────────────────────────────────────
ws2 = wb.create_sheet(title="Campaign Summary")
summary_data = [
    ["Metric", "Value"],
    ["Total Rows", len(rows)],
    ["Total Days", 30],
    ["Campaign Start", "2026-05-01"],
    ["Campaign End", "2026-05-30"],
    ["Platforms", "Instagram, Facebook, Blog"],
    ["", ""],
    ["Content Type", "Count"],
]

type_counts = {}
for r in rows:
    ct = r[5]
    type_counts[ct] = type_counts.get(ct, 0) + 1

for ct, count in sorted(type_counts.items()):
    summary_data.append([ct, count])

summary_data += [
    ["", ""],
    ["Week", "Row Count"],
]
week_counts = {}
for r in rows:
    wk = r[2]
    week_counts[wk] = week_counts.get(wk, 0) + 1
for wk, count in sorted(week_counts.items()):
    summary_data.append([wk, count])

for s_row in summary_data:
    ws2.append(s_row)

# style summary header rows
for row_idx2, s_row in enumerate(summary_data, start=1):
    for col_idx2 in range(1, 3):
        cell = ws2.cell(row=row_idx2, column=col_idx2)
        cell.alignment = left_align
        cell.border = thin_border
        if s_row and s_row[0] in ("Metric", "Content Type", "Week"):
            cell.fill   = header_fill
            cell.font   = header_font

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 18

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")
print(f"Total data rows: {len(rows)}")
print(f"Content type breakdown: {type_counts}")
